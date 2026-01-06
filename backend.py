"""
Backend Control CxC v4 - Multi-cliente
Conexión con Google Sheets + Generación de Reportes

Instrucciones:
1. pip install flask flask-cors gspread google-auth openpyxl reportlab bcrypt PyJWT
2. Colocar credentials.json en la misma carpeta
3. python backend.py
4. Abrir el HTML en el navegador
"""

from flask import Flask, request, jsonify, send_file, make_response
from flask_cors import CORS
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime, timedelta
import os
import io
import json
import secrets
import hashlib
from functools import wraps

# Para autenticación
import jwt
import bcrypt

# Para reportes Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Para reportes PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

app = Flask(__name__)
CORS(app, supports_credentials=True)

# =====================
# Configuración
# =====================
CREDENTIALS_FILE = "credentials.json"
NEGOCIOS_FILE = "negocios.json"

# Para despliegue en nube: credenciales desde variable de entorno
GOOGLE_CREDENTIALS_JSON = os.environ.get('GOOGLE_CREDENTIALS_JSON')

# Clave secreta para JWT (generar una única y guardar en variable de entorno)
JWT_SECRET = os.environ.get('JWT_SECRET', secrets.token_hex(32))
JWT_EXPIRATION_HOURS = int(os.environ.get('JWT_EXPIRATION_HOURS', '8'))

# Negocio activo (se cambia dinámicamente)
current_sheet_id = None

# Sheet maestro para guardar la lista de negocios (persistente)
# Este Sheet ID debe configurarse como variable de entorno MASTER_SHEET_ID
MASTER_SHEET_ID = os.environ.get('MASTER_SHEET_ID', '')

# Headers para la hoja de Usuarios
HEADERS_USUARIOS = ['ID', 'Usuario', 'PasswordHash', 'Nombre', 'Email', 'Rol', 'Activo', 'UltimoAcceso', 'FechaCreacion']

# =====================
# AUTENTICACIÓN
# =====================
def hash_password(password):
    """Hashea una contraseña con bcrypt"""
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password, hashed):
    """Verifica una contraseña contra su hash"""
    try:
        return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))
    except:
        return False

def generate_token(user_data):
    """Genera un JWT token"""
    payload = {
        'user_id': user_data['id'],
        'usuario': user_data['usuario'],
        'nombre': user_data['nombre'],
        'rol': user_data.get('rol', 'usuario'),
        'exp': datetime.utcnow() + timedelta(hours=JWT_EXPIRATION_HOURS),
        'iat': datetime.utcnow()
    }
    return jwt.encode(payload, JWT_SECRET, algorithm='HS256')

def verify_token(token):
    """Verifica y decodifica un JWT token"""
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
        return payload
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None

def auth_required(f):
    """Decorator para proteger endpoints que requieren autenticación"""
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        
        # Buscar token en header Authorization
        auth_header = request.headers.get('Authorization')
        if auth_header and auth_header.startswith('Bearer '):
            token = auth_header.split(' ')[1]
        
        # También buscar en cookies
        if not token:
            token = request.cookies.get('auth_token')
        
        if not token:
            return jsonify({'success': False, 'error': 'Token de autenticación requerido'}), 401
        
        payload = verify_token(token)
        if not payload:
            return jsonify({'success': False, 'error': 'Token inválido o expirado'}), 401
        
        # Agregar datos del usuario al request
        request.user = payload
        return f(*args, **kwargs)
    
    return decorated

def admin_required(f):
    """Decorator para endpoints que requieren rol de administrador"""
    @wraps(f)
    @auth_required
    def decorated(*args, **kwargs):
        if request.user.get('rol') != 'admin':
            return jsonify({'success': False, 'error': 'Se requieren permisos de administrador'}), 403
        return f(*args, **kwargs)
    return decorated

def get_usuarios_worksheet():
    """Obtiene la hoja de usuarios del Master Sheet"""
    master = get_master_sheet()
    if not master:
        return None
    
    try:
        ws = master.worksheet('Usuarios')
    except gspread.WorksheetNotFound:
        ws = master.add_worksheet(title='Usuarios', rows=100, cols=len(HEADERS_USUARIOS))
        ws.append_row(HEADERS_USUARIOS)
    return ws

# =====================
# ENDPOINTS DE AUTENTICACIÓN
# =====================
@app.route('/api/auth/login', methods=['POST'])
def login():
    """Inicia sesión y retorna un JWT token"""
    try:
        data = request.json
        usuario = data.get('usuario', '').strip().lower()
        password = data.get('password', '')
        
        if not usuario or not password:
            return jsonify({'success': False, 'error': 'Usuario y contraseña requeridos'}), 400
        
        ws = get_usuarios_worksheet()
        
        # Obtener usuarios existentes
        usuarios = []
        if ws:
            try:
                usuarios = ws.get_all_records()
            except:
                usuarios = []
        
        # Si no hay usuarios registrados, permitir admin/admin por defecto
        if len(usuarios) == 0:
            if usuario == 'admin' and password == 'admin':
                token = generate_token({
                    'id': '1',
                    'usuario': 'admin',
                    'nombre': 'Administrador',
                    'rol': 'admin'
                })
                response = make_response(jsonify({
                    'success': True,
                    'token': token,
                    'user': {
                        'id': '1',
                        'usuario': 'admin',
                        'nombre': 'Administrador',
                        'rol': 'admin'
                    },
                    'mensaje': '⚠️ Usando credenciales por defecto. Creá un usuario desde Configuración.'
                }))
                response.set_cookie('auth_token', token, httponly=True, secure=True, samesite='None', max_age=JWT_EXPIRATION_HOURS*3600)
                return response
            else:
                return jsonify({'success': False, 'error': 'Credenciales inválidas. Use admin/admin para el primer acceso.'}), 401
        
        # Buscar usuario en la lista
        user = None
        for u in usuarios:
            if u.get('Usuario', '').lower() == usuario and str(u.get('Activo', 'TRUE')).upper() == 'TRUE':
                user = u
                break
        
        if not user:
            return jsonify({'success': False, 'error': 'Usuario no encontrado o inactivo'}), 401
        
        # Verificar contraseña
        if not verify_password(password, user.get('PasswordHash', '')):
            return jsonify({'success': False, 'error': 'Contraseña incorrecta'}), 401
        
        # Actualizar último acceso
        try:
            cell = ws.find(user.get('Usuario'))
            if cell:
                ws.update_cell(cell.row, 8, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        except:
            pass
        
        # Generar token
        user_data = {
            'id': str(user.get('ID', '')),
            'usuario': user.get('Usuario', ''),
            'nombre': user.get('Nombre', ''),
            'rol': user.get('Rol', 'usuario')
        }
        token = generate_token(user_data)
        
        response = make_response(jsonify({
            'success': True,
            'token': token,
            'user': user_data
        }))
        response.set_cookie('auth_token', token, httponly=True, secure=True, samesite='None', max_age=JWT_EXPIRATION_HOURS*3600)
        return response
        
    except Exception as e:
        print(f"Error en login: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    """Cierra la sesión"""
    response = make_response(jsonify({'success': True}))
    response.delete_cookie('auth_token')
    return response

@app.route('/api/auth/verify', methods=['GET'])
def verify_auth():
    """Verifica si el token actual es válido"""
    token = None
    
    auth_header = request.headers.get('Authorization')
    if auth_header and auth_header.startswith('Bearer '):
        token = auth_header.split(' ')[1]
    
    if not token:
        token = request.cookies.get('auth_token')
    
    if not token:
        return jsonify({'success': False, 'authenticated': False})
    
    payload = verify_token(token)
    if not payload:
        return jsonify({'success': False, 'authenticated': False})
    
    return jsonify({
        'success': True,
        'authenticated': True,
        'user': {
            'id': payload.get('user_id'),
            'usuario': payload.get('usuario'),
            'nombre': payload.get('nombre'),
            'rol': payload.get('rol')
        }
    })

@app.route('/api/auth/cambiar-password', methods=['POST'])
@auth_required
def cambiar_password():
    """Cambia la contraseña del usuario actual"""
    try:
        data = request.json
        password_actual = data.get('passwordActual', '')
        password_nueva = data.get('passwordNueva', '')
        
        if not password_actual or not password_nueva:
            return jsonify({'success': False, 'error': 'Ambas contraseñas son requeridas'}), 400
        
        if len(password_nueva) < 6:
            return jsonify({'success': False, 'error': 'La nueva contraseña debe tener al menos 6 caracteres'}), 400
        
        ws = get_usuarios_worksheet()
        if not ws:
            return jsonify({'success': False, 'error': 'Sistema de usuarios no configurado'}), 500
        
        usuarios = ws.get_all_records()
        user = None
        row_num = None
        
        for i, u in enumerate(usuarios):
            if str(u.get('ID', '')) == request.user.get('user_id'):
                user = u
                row_num = i + 2  # +2 por header y índice base 0
                break
        
        if not user:
            return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404
        
        # Verificar contraseña actual
        if not verify_password(password_actual, user.get('PasswordHash', '')):
            return jsonify({'success': False, 'error': 'Contraseña actual incorrecta'}), 401
        
        # Actualizar contraseña
        new_hash = hash_password(password_nueva)
        ws.update_cell(row_num, 3, new_hash)
        
        return jsonify({'success': True, 'mensaje': 'Contraseña actualizada correctamente'})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# =====================
# GESTIÓN DE USUARIOS (Solo Admin)
# =====================
@app.route('/api/usuarios', methods=['GET'])
@admin_required
def get_usuarios():
    """Lista todos los usuarios"""
    try:
        ws = get_usuarios_worksheet()
        if not ws:
            return jsonify({'success': True, 'data': []})
        
        try:
            usuarios = ws.get_all_records()
        except:
            usuarios = []
        
        # No enviar el hash de contraseña
        usuarios_safe = []
        for u in usuarios:
            usuarios_safe.append({
                'id': str(u.get('ID', '')),
                'usuario': u.get('Usuario', ''),
                'nombre': u.get('Nombre', ''),
                'email': u.get('Email', ''),
                'rol': u.get('Rol', 'usuario'),
                'activo': u.get('Activo', 'TRUE') == 'TRUE',
                'ultimoAcceso': u.get('UltimoAcceso', ''),
                'fechaCreacion': u.get('FechaCreacion', '')
            })
        
        return jsonify({'success': True, 'data': usuarios_safe})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/usuarios', methods=['POST'])
@admin_required
def crear_usuario():
    """Crea un nuevo usuario"""
    try:
        data = request.json
        usuario = data.get('usuario', '').strip().lower()
        password = data.get('password', '')
        nombre = data.get('nombre', '')
        email = data.get('email', '')
        rol = data.get('rol', 'usuario')
        
        if not usuario or not password or not nombre:
            return jsonify({'success': False, 'error': 'Usuario, contraseña y nombre son requeridos'}), 400
        
        if len(password) < 6:
            return jsonify({'success': False, 'error': 'La contraseña debe tener al menos 6 caracteres'}), 400
        
        if rol not in ['admin', 'usuario', 'lectura']:
            return jsonify({'success': False, 'error': 'Rol inválido'}), 400
        
        ws = get_usuarios_worksheet()
        if not ws:
            return jsonify({'success': False, 'error': 'Sistema de usuarios no configurado'}), 500
        
        # Verificar que el usuario no exista
        try:
            usuarios = ws.get_all_records()
        except:
            usuarios = []
        
        for u in usuarios:
            if u.get('Usuario', '').lower() == usuario:
                return jsonify({'success': False, 'error': 'El usuario ya existe'}), 400
        
        # Crear usuario
        user_id = datetime.now().strftime('%Y%m%d%H%M%S')
        password_hash = hash_password(password)
        
        ws.append_row([
            user_id,
            usuario,
            password_hash,
            nombre,
            email,
            rol,
            'TRUE',
            '',
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ])
        
        return jsonify({
            'success': True,
            'usuario': {
                'id': user_id,
                'usuario': usuario,
                'nombre': nombre,
                'rol': rol
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/usuarios/<user_id>', methods=['PUT'])
@admin_required
def actualizar_usuario(user_id):
    """Actualiza un usuario existente"""
    try:
        data = request.json
        
        ws = get_usuarios_worksheet()
        if not ws:
            return jsonify({'success': False, 'error': 'Sistema no configurado'}), 500
        
        usuarios = ws.get_all_records()
        row_num = None
        
        for i, u in enumerate(usuarios):
            if str(u.get('ID', '')) == user_id:
                row_num = i + 2
                break
        
        if not row_num:
            return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404
        
        # Actualizar campos
        if 'nombre' in data:
            ws.update_cell(row_num, 4, data['nombre'])
        if 'email' in data:
            ws.update_cell(row_num, 5, data['email'])
        if 'rol' in data and data['rol'] in ['admin', 'usuario', 'lectura']:
            ws.update_cell(row_num, 6, data['rol'])
        if 'activo' in data:
            ws.update_cell(row_num, 7, 'TRUE' if data['activo'] else 'FALSE')
        if 'password' in data and len(data['password']) >= 6:
            ws.update_cell(row_num, 3, hash_password(data['password']))
        
        return jsonify({'success': True})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/usuarios/<user_id>', methods=['DELETE'])
@admin_required
def eliminar_usuario(user_id):
    """Elimina (desactiva) un usuario"""
    try:
        # No permitir eliminar el propio usuario
        if request.user.get('user_id') == user_id:
            return jsonify({'success': False, 'error': 'No podés eliminar tu propio usuario'}), 400
        
        ws = get_usuarios_worksheet()
        if not ws:
            return jsonify({'success': False, 'error': 'Sistema no configurado'}), 500
        
        usuarios = ws.get_all_records()
        row_num = None
        
        for i, u in enumerate(usuarios):
            if str(u.get('ID', '')) == user_id:
                row_num = i + 2
                break
        
        if not row_num:
            return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404
        
        # Solo desactivar, no eliminar físicamente
        ws.update_cell(row_num, 7, 'FALSE')
        
        return jsonify({'success': True})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/auth/setup', methods=['POST'])
def setup_admin():
    """Crea el usuario admin inicial (solo si no existe ningún usuario)"""
    try:
        ws = get_usuarios_worksheet()
        if not ws:
            return jsonify({'success': False, 'error': 'Configure MASTER_SHEET_ID primero'}), 500
        
        try:
            usuarios = ws.get_all_records()
        except:
            usuarios = []
        
        if len(usuarios) > 0:
            return jsonify({'success': False, 'error': 'Ya existe al menos un usuario'}), 400
        
        data = request.json
        password = data.get('password', '')
        
        if len(password) < 6:
            return jsonify({'success': False, 'error': 'La contraseña debe tener al menos 6 caracteres'}), 400
        
        # Crear usuario admin
        user_id = datetime.now().strftime('%Y%m%d%H%M%S')
        password_hash = hash_password(password)
        
        ws.append_row([
            user_id,
            'admin',
            password_hash,
            'Administrador',
            '',
            'admin',
            'TRUE',
            '',
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ])
        
        return jsonify({
            'success': True,
            'mensaje': 'Usuario admin creado exitosamente',
            'usuario': 'admin'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

def get_master_sheet():
    """Obtiene conexión al Sheet maestro"""
    if not MASTER_SHEET_ID:
        return None
    
    scopes = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive'
    ]
    
    if GOOGLE_CREDENTIALS_JSON:
        creds_dict = json.loads(GOOGLE_CREDENTIALS_JSON)
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    elif os.path.exists(CREDENTIALS_FILE):
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scopes)
    else:
        return None
    
    client = gspread.authorize(creds)
    return client.open_by_key(MASTER_SHEET_ID)

def load_negocios():
    """Carga la lista de negocios desde Sheet maestro, variable de entorno o archivo local"""
    # 1. Primero intentar desde Sheet maestro (más persistente)
    try:
        master = get_master_sheet()
        if master:
            try:
                ws = master.worksheet('Negocios')
                records = ws.get_all_records()
                negocios = []
                for r in records:
                    negocios.append({
                        'id': str(r.get('ID', '')),
                        'nombre': r.get('Nombre', ''),
                        'sheetId': r.get('SheetID', ''),
                        'descripcion': r.get('Descripcion', ''),
                        'activo': r.get('Activo', 'TRUE') == 'TRUE'
                    })
                return negocios
            except:
                pass
    except:
        pass
    
    # 2. Intentar variable de entorno (para la nube sin Sheet maestro)
    negocios_env = os.environ.get('NEGOCIOS_JSON')
    if negocios_env:
        return json.loads(negocios_env)
    
    # 3. Si no, usar archivo local
    if os.path.exists(NEGOCIOS_FILE):
        with open(NEGOCIOS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def save_negocios(negocios):
    """Guarda la lista de negocios en Sheet maestro, variable de entorno o archivo local"""
    # 1. Guardar en Sheet maestro si está configurado
    try:
        master = get_master_sheet()
        if master:
            try:
                ws = master.worksheet('Negocios')
            except:
                ws = master.add_worksheet(title='Negocios', rows=100, cols=5)
                ws.append_row(['ID', 'Nombre', 'SheetID', 'Descripcion', 'Activo'])
            
            # Limpiar y reescribir
            ws.clear()
            ws.append_row(['ID', 'Nombre', 'SheetID', 'Descripcion', 'Activo'])
            
            for n in negocios:
                ws.append_row([
                    n.get('id', ''),
                    n.get('nombre', ''),
                    n.get('sheetId', ''),
                    n.get('descripcion', ''),
                    'TRUE' if n.get('activo', True) else 'FALSE'
                ])
            return
    except Exception as e:
        print(f"Error guardando en Sheet maestro: {e}")
    
    # 2. Guardar en variable de entorno (memoria, se pierde al reiniciar)
    if os.environ.get('NEGOCIOS_JSON') is not None:
        os.environ['NEGOCIOS_JSON'] = json.dumps(negocios, ensure_ascii=False)
    else:
        # 3. Guardar en archivo local
        with open(NEGOCIOS_FILE, 'w', encoding='utf-8') as f:
            json.dump(negocios, f, ensure_ascii=False, indent=2)

def get_current_sheet_id():
    """Obtiene el Sheet ID del negocio activo"""
    global current_sheet_id
    if current_sheet_id:
        return current_sheet_id
    # Si no hay negocio seleccionado, usar el primero
    negocios = load_negocios()
    if negocios:
        current_sheet_id = negocios[0].get('sheetId')
        return current_sheet_id
    return None

# =====================
# Helper: Parsear números (maneja , como decimal)
# =====================
def parse_number(value):
    """Convierte un valor a float, manejando diferentes formatos"""
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    
    # Si es string, extraer SOLO dígitos, punto, coma y guion
    s = str(value)
    limpio = ''
    for c in s:
        if c.isdigit() or c in '.,-':
            limpio += c
    
    s = limpio.strip()
    
    if not s or s == '-':
        return 0.0
    
    # Determinar separador decimal
    if ',' in s and '.' in s:
        last_comma = s.rfind(',')
        last_dot = s.rfind('.')
        if last_comma > last_dot:
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        s = s.replace(',', '.')
    
    try:
        return float(s)
    except:
        return 0.0

# =====================
# Conexión Google Sheets
# =====================
def get_sheet(sheet_id=None):
    """Obtiene conexión al Google Sheet"""
    scopes = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive'
    ]
    
    # Intentar credenciales desde variable de entorno (nube)
    if GOOGLE_CREDENTIALS_JSON:
        creds_dict = json.loads(GOOGLE_CREDENTIALS_JSON)
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    elif os.path.exists(CREDENTIALS_FILE):
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scopes)
    else:
        raise Exception("No se encontraron credenciales de Google")
    
    client = gspread.authorize(creds)
    
    # Usar el sheet_id pasado, o el actual, o el primero disponible
    target_sheet_id = sheet_id or get_current_sheet_id()
    if not target_sheet_id:
        raise Exception("No hay negocio configurado. Agregá uno en Configuración.")
    
    return client.open_by_key(target_sheet_id)

def get_or_create_worksheet(sheet, name, headers):
    try:
        ws = sheet.worksheet(name)
        # Verificar si la hoja tiene encabezados, si no, agregarlos
        try:
            first_row = ws.row_values(1)
            if not first_row or len(first_row) == 0:
                ws.append_row(headers)
        except:
            ws.append_row(headers)
    except gspread.WorksheetNotFound:
        ws = sheet.add_worksheet(title=name, rows=1000, cols=len(headers))
        ws.append_row(headers)
    return ws

# Headers para las hojas
HEADERS_CLIENTES = ['ID', 'Identificacion', 'Nombre', 'DiasCredito', 'Activo', 'FechaCreacion']
HEADERS_FACTURAS = ['ID', 'Consecutivo', 'Fecha', 'ClienteID', 'ClienteNombre', 'CedulaCliente',
                    'TotalFactura', 'CORFOGA', 'OtrosRebajos', 'MontoCobrar', 
                    'FechaVencimiento', 'Pagado', 'FechaPago', 'TipoProducto', 
                    'OrdenCompra', 'Notas', 'TipoDocumento', 'DocumentoRelacionado', 'Estado',
                    'TotalAbonado', 'SaldoPendiente']
HEADERS_ABONOS = ['ID', 'FacturaID', 'Consecutivo', 'Fecha', 'Monto', 'MetodoPago', 'Referencia', 'Notas']

def detectar_tipo_documento(consecutivo):
    """Detecta si es Factura o Nota de Crédito basado en el consecutivo"""
    cons = str(consecutivo)
    # El tipo de documento está en posición 6-7 (después de 10000X)
    # 01 = Factura, 03 = Nota de Crédito
    if len(cons) >= 8:
        tipo_codigo = cons[6:8]
        if tipo_codigo == '03':
            return 'NC'
    return 'FAC'

# =====================
# NEGOCIOS (Multi-cliente)
# =====================
@app.route('/api/negocios', methods=['GET'])
@auth_required
def get_negocios():
    """Obtiene la lista de negocios configurados"""
    try:
        negocios = load_negocios()
        # Marcar cuál es el activo
        for n in negocios:
            n['activo'] = (n.get('sheetId') == current_sheet_id)
        return jsonify({'success': True, 'data': negocios})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/negocios/debug', methods=['GET'])
@auth_required
def debug_negocios():
    """Endpoint de diagnóstico para verificar configuración de negocios"""
    debug_info = {
        'MASTER_SHEET_ID_configurado': bool(MASTER_SHEET_ID),
        'MASTER_SHEET_ID_valor': MASTER_SHEET_ID[:15] + '...' if MASTER_SHEET_ID else 'NO CONFIGURADO',
        'GOOGLE_CREDENTIALS_JSON_configurado': bool(GOOGLE_CREDENTIALS_JSON),
        'negocios_cargados': 0,
        'master_sheet_conexion': False,
        'hoja_negocios_existe': False,
        'error': None
    }
    
    try:
        negocios = load_negocios()
        debug_info['negocios_cargados'] = len(negocios)
    except Exception as e:
        debug_info['error_carga'] = str(e)
    
    try:
        master = get_master_sheet()
        if master:
            debug_info['master_sheet_conexion'] = True
            debug_info['master_sheet_titulo'] = master.title
            try:
                ws = master.worksheet('Negocios')
                debug_info['hoja_negocios_existe'] = True
                debug_info['filas_en_hoja'] = len(ws.get_all_records())
            except Exception as e:
                debug_info['hoja_negocios_existe'] = False
                debug_info['error_hoja'] = str(e)
        else:
            debug_info['error'] = 'get_master_sheet() retornó None'
    except Exception as e:
        debug_info['error'] = str(e)
    
    return jsonify(debug_info)

@app.route('/api/negocios', methods=['POST'])
@auth_required
def add_negocio():
    """Agrega un nuevo negocio"""
    try:
        data = request.json
        negocios = load_negocios()
        
        nuevo_negocio = {
            'id': datetime.now().strftime('%Y%m%d%H%M%S'),
            'nombre': data.get('nombre', ''),
            'sheetId': data.get('sheetId', ''),
            'descripcion': data.get('descripcion', '')
        }
        
        negocios.append(nuevo_negocio)
        save_negocios(negocios)
        
        # Si es el primer negocio, activarlo automáticamente
        global current_sheet_id
        if len(negocios) == 1:
            current_sheet_id = nuevo_negocio['sheetId']
        
        return jsonify({'success': True, 'negocio': nuevo_negocio})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/negocios/<negocio_id>', methods=['PUT'])
@auth_required
def update_negocio(negocio_id):
    """Actualiza un negocio existente"""
    try:
        data = request.json
        negocios = load_negocios()
        
        for n in negocios:
            if n['id'] == negocio_id:
                n['nombre'] = data.get('nombre', n['nombre'])
                n['sheetId'] = data.get('sheetId', n['sheetId'])
                n['descripcion'] = data.get('descripcion', n['descripcion'])
                break
        
        save_negocios(negocios)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/negocios/<negocio_id>', methods=['DELETE'])
@admin_required
def delete_negocio(negocio_id):
    """Elimina un negocio"""
    try:
        negocios = load_negocios()
        negocios = [n for n in negocios if n['id'] != negocio_id]
        save_negocios(negocios)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/negocios/activar/<negocio_id>', methods=['POST'])
@auth_required
def activar_negocio(negocio_id):
    """Activa un negocio (lo hace el actual)"""
    try:
        global current_sheet_id
        negocios = load_negocios()
        
        for n in negocios:
            if n['id'] == negocio_id:
                current_sheet_id = n['sheetId']
                return jsonify({'success': True, 'sheetId': current_sheet_id, 'nombre': n['nombre']})
        
        return jsonify({'success': False, 'error': 'Negocio no encontrado'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/negocios/actual', methods=['GET'])
@auth_required
def get_negocio_actual():
    """Obtiene el negocio actualmente activo"""
    global current_sheet_id
    try:
        negocios = load_negocios()
        for n in negocios:
            if n.get('sheetId') == current_sheet_id:
                return jsonify({'success': True, 'negocio': n})
        
        # Si no hay ninguno activo pero hay negocios, activar el primero
        if negocios:
            current_sheet_id = negocios[0]['sheetId']
            return jsonify({'success': True, 'negocio': negocios[0]})
        
        return jsonify({'success': True, 'negocio': None})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# =====================
# CLIENTES
# =====================
@app.route('/api/clientes', methods=['GET'])
@auth_required
def get_clientes():
    try:
        sheet = get_sheet()
        ws = get_or_create_worksheet(sheet, 'Clientes', HEADERS_CLIENTES)
        
        try:
            records = ws.get_all_records()
        except Exception as e:
            # Si falla get_all_records (hoja vacía), retornar lista vacía
            print(f"Advertencia al leer clientes: {e}")
            records = []
        
        # Convertir a formato esperado por el frontend
        clientes = []
        for r in records:
            clientes.append({
                'id': str(r.get('ID', '')),
                'identificacion': str(r.get('Identificacion', '')),
                'nombre': r.get('Nombre', ''),
                'diasVencimiento': int(r.get('DiasCredito', 8) or 8),
                'activo': r.get('Activo', 'TRUE') == 'TRUE'
            })
        
        return jsonify({'success': True, 'data': clientes})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/clientes', methods=['POST'])
@auth_required
def add_cliente():
    try:
        data = request.json
        sheet = get_sheet()
        ws = get_or_create_worksheet(sheet, 'Clientes', HEADERS_CLIENTES)
        
        cliente_id = datetime.now().strftime('%Y%m%d%H%M%S%f')
        row = [
            cliente_id,
            data.get('identificacion', ''),
            data.get('nombre', ''),
            data.get('diasVencimiento', 8),
            'TRUE',
            datetime.now().isoformat()
        ]
        ws.append_row(row)
        
        return jsonify({'success': True, 'id': cliente_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/clientes/<cliente_id>', methods=['PUT'])
@auth_required
def update_cliente(cliente_id):
    try:
        data = request.json
        sheet = get_sheet()
        ws = sheet.worksheet('Clientes')
        
        # Buscar fila del cliente
        cell = ws.find(cliente_id)
        if cell:
            row = cell.row
            if 'nombre' in data:
                ws.update_cell(row, 3, data['nombre'])
            if 'identificacion' in data:
                ws.update_cell(row, 2, data['identificacion'])
            if 'diasVencimiento' in data:
                ws.update_cell(row, 4, data['diasVencimiento'])
            if 'activo' in data:
                ws.update_cell(row, 5, 'TRUE' if data['activo'] else 'FALSE')
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/clientes/<cliente_id>/toggle', methods=['POST'])
@auth_required
def toggle_cliente(cliente_id):
    try:
        sheet = get_sheet()
        ws = sheet.worksheet('Clientes')
        
        cell = ws.find(cliente_id)
        if cell:
            row = cell.row
            current = ws.cell(row, 5).value
            new_value = 'FALSE' if current == 'TRUE' else 'TRUE'
            ws.update_cell(row, 5, new_value)
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# =====================
# FACTURAS
# =====================
@app.route('/api/facturas', methods=['GET'])
@auth_required
def get_facturas():
    try:
        sheet = get_sheet()
        ws = get_or_create_worksheet(sheet, 'Facturas', HEADERS_FACTURAS)
        
        try:
            records = ws.get_all_records()
        except Exception as e:
            # Si falla get_all_records (hoja vacía o sin datos), retornar lista vacía
            print(f"Advertencia al leer facturas: {e}")
            records = []
        
        facturas = []
        for r in records:
            try:
                consecutivo = str(r.get('Consecutivo', ''))
                tipo_doc = r.get('TipoDocumento', '') or detectar_tipo_documento(consecutivo)
                estado = r.get('Estado', '')
                
                # Determinar estado si no existe
                if not estado:
                    if r.get('Pagado', 'FALSE') == 'TRUE':
                        estado = 'Pagado'
                    else:
                        estado = 'Pendiente'
                
                facturas.append({
                    'id': str(r.get('ID', '')),
                    'consecutivo': consecutivo,
                    'fecha': r.get('Fecha', ''),
                    'clienteId': str(r.get('ClienteID', '')),
                    'clienteNombre': r.get('ClienteNombre', ''),
                    'cedulaCliente': str(r.get('CedulaCliente', '')),
                    'totalFactura': parse_number(r.get('TotalFactura')),
                    'corfoga': parse_number(r.get('CORFOGA')),
                    'otrosRebajos': parse_number(r.get('OtrosRebajos')),
                    'montoCobrar': parse_number(r.get('MontoCobrar')),
                    'fechaVencimiento': r.get('FechaVencimiento', ''),
                    'pagado': estado in ['Pagado', 'Compensado'],
                    'fechaPago': r.get('FechaPago', ''),
                    'tipoProducto': r.get('TipoProducto', ''),
                    'ordenCompra': r.get('OrdenCompra', ''),
                    'notas': r.get('Notas', ''),
                    'tipoDocumento': tipo_doc,
                    'documentoRelacionado': str(r.get('DocumentoRelacionado', '') or ''),
                    'estado': estado,
                    'totalAbonado': parse_number(r.get('TotalAbonado', 0)),
                    'saldoPendiente': parse_number(r.get('SaldoPendiente', 0)) or parse_number(r.get('MontoCobrar', 0)) - parse_number(r.get('TotalAbonado', 0))
                })
            except Exception as row_error:
                print(f"Error procesando fila: {row_error}")
                continue
        
        return jsonify({'success': True, 'data': facturas})
    except Exception as e:
        print(f"Error en get_facturas: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/facturas', methods=['POST'])
@auth_required
def add_factura():
    try:
        data = request.json
        sheet = get_sheet()
        ws = get_or_create_worksheet(sheet, 'Facturas', HEADERS_FACTURAS)
        
        factura_id = datetime.now().strftime('%Y%m%d%H%M%S%f')
        total = float(data.get('totalFactura', 0))
        corfoga = float(data.get('corfoga', 0))
        otros = float(data.get('otrosRebajos', 0))
        monto_cobrar = total - corfoga - otros
        consecutivo = str(data.get('consecutivo', ''))
        tipo_doc = data.get('tipoDocumento', '') or detectar_tipo_documento(consecutivo)
        
        row = [
            factura_id,
            consecutivo,
            data.get('fecha', ''),
            data.get('clienteId', ''),
            data.get('clienteNombre', ''),
            data.get('cedulaCliente', ''),
            total,
            corfoga,
            otros,
            monto_cobrar,
            data.get('fechaVencimiento', ''),
            'FALSE',
            '',
            data.get('tipoProducto', ''),
            data.get('ordenCompra', ''),
            data.get('notas', ''),
            tipo_doc,
            data.get('documentoRelacionado', ''),
            'Pendiente'
        ]
        ws.append_row(row)
        
        return jsonify({'success': True, 'id': factura_id, 'tipoDocumento': tipo_doc})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/facturas/batch', methods=['POST'])
@auth_required
def add_facturas_batch():
    try:
        facturas = request.json.get('facturas', [])
        sheet = get_sheet()
        ws = get_or_create_worksheet(sheet, 'Facturas', HEADERS_FACTURAS)
        
        # Obtener consecutivos existentes para evitar duplicados
        # Normalizar quitando ceros iniciales para comparación
        try:
            existing_records = ws.get_all_records()
        except Exception as e:
            print(f"Advertencia al leer facturas existentes: {e}")
            existing_records = []
        
        existing_consecutivos = set(str(r.get('Consecutivo', '')).strip().lstrip('0') for r in existing_records)
        
        rows = []
        notas_credito = []
        duplicados = []
        
        for i, data in enumerate(facturas):
            consecutivo = str(data.get('consecutivo', '')).strip()
            consecutivo_normalizado = consecutivo.lstrip('0')
            
            # Verificar si ya existe
            if consecutivo_normalizado in existing_consecutivos:
                duplicados.append(consecutivo)
                continue  # Saltar esta factura
            
            factura_id = datetime.now().strftime('%Y%m%d%H%M%S') + str(i).zfill(4)
            total = float(data.get('totalFactura', 0))
            corfoga = float(data.get('corfoga', 0))
            otros = float(data.get('otrosRebajos', 0))
            monto_cobrar = total - corfoga - otros
            tipo_doc = detectar_tipo_documento(consecutivo)
            
            row_data = [
                factura_id,
                consecutivo,
                data.get('fecha', ''),
                data.get('clienteId', ''),
                data.get('clienteNombre', ''),
                data.get('cedulaCliente', ''),
                total,
                corfoga,
                otros,
                monto_cobrar,
                data.get('fechaVencimiento', ''),
                'FALSE',
                '',
                data.get('tipoProducto', ''),
                data.get('ordenCompra', ''),
                data.get('notas', ''),
                tipo_doc,
                '',  # DocumentoRelacionado - se llena después si es NC
                'Pendiente'
            ]
            rows.append(row_data)
            
            if tipo_doc == 'NC':
                notas_credito.append({
                    'id': factura_id,
                    'consecutivo': consecutivo,
                    'clienteId': data.get('clienteId', ''),
                    'monto': monto_cobrar
                })
        
        if rows:
            ws.append_rows(rows)
        
        return jsonify({
            'success': True, 
            'count': len(rows),
            'duplicados': len(duplicados),
            'duplicadosLista': duplicados[:10],  # Máximo 10 para no saturar
            'notasCredito': notas_credito
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/factura/<factura_id>', methods=['GET'])
@auth_required
def get_factura_by_id(factura_id):
    """Obtiene una factura por su ID"""
    try:
        sheet = get_sheet()
        ws = get_or_create_worksheet(sheet, 'Facturas', HEADERS_FACTURAS)
        
        try:
            records = ws.get_all_records()
        except:
            records = []
        
        # Buscar por ID
        for r in records:
            if str(r.get('ID', '')) == str(factura_id):
                return jsonify({'success': True, 'factura': r})
        
        # Si no se encuentra por ID, intentar por Consecutivo
        for r in records:
            if str(r.get('Consecutivo', '')) == str(factura_id):
                return jsonify({'success': True, 'factura': r})
        
        return jsonify({'success': False, 'error': 'Factura no encontrada'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/facturas/<factura_id>', methods=['PUT'])
@auth_required
def update_factura(factura_id):
    try:
        data = request.json
        sheet = get_sheet()
        ws = sheet.worksheet('Facturas')
        
        cell = ws.find(factura_id)
        if cell:
            row = cell.row
            
            # Obtener valores actuales para recálculo
            total = float(ws.cell(row, 7).value or 0)
            
            if 'totalFactura' in data:
                total = float(data['totalFactura'])
                ws.update_cell(row, 7, total)
            
            corfoga = float(data.get('corfoga', ws.cell(row, 8).value or 0))
            otros = float(data.get('otrosRebajos', ws.cell(row, 9).value or 0))
            
            if 'corfoga' in data:
                ws.update_cell(row, 8, corfoga)
            if 'otrosRebajos' in data:
                ws.update_cell(row, 9, otros)
            
            # Recalcular monto a cobrar
            monto_cobrar = total - corfoga - otros
            ws.update_cell(row, 10, monto_cobrar)
            
            if 'tipoProducto' in data:
                ws.update_cell(row, 14, data['tipoProducto'])
            if 'ordenCompra' in data:
                ws.update_cell(row, 15, data['ordenCompra'])
            if 'notas' in data:
                ws.update_cell(row, 16, data['notas'])
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/facturas/<factura_id>/pago', methods=['POST'])
@auth_required
def registrar_pago(factura_id):
    try:
        data = request.json
        sheet = get_sheet()
        ws = sheet.worksheet('Facturas')
        
        cell = ws.find(factura_id)
        if cell:
            row = cell.row
            monto_parcial = data.get('montoParcial')
            
            if monto_parcial:
                # Pago parcial - actualizar MontoCobrar
                monto_actual = parse_number(ws.cell(row, 10).value)  # MontoCobrar
                nuevo_monto = monto_actual - float(monto_parcial)
                ws.update_cell(row, 10, nuevo_monto)  # Actualizar MontoCobrar
                # NO marcar como pagado, sigue pendiente
            else:
                # Pago total
                ws.update_cell(row, 12, 'TRUE')  # Pagado
                ws.update_cell(row, 19, 'Pagado')  # Estado
            
            ws.update_cell(row, 13, data.get('fechaPago', datetime.now().strftime('%Y-%m-%d')))
            
            # Agregar nota de pago
            notas_actuales = ws.cell(row, 16).value or ''
            nueva_nota = data.get('notas', '')
            if nueva_nota:
                ws.update_cell(row, 16, f"{notas_actuales} | Pago: {nueva_nota}" if notas_actuales else f"Pago: {nueva_nota}")
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/facturas/compensar', methods=['POST'])
def compensar_documentos():
    """Compensa una Nota de Crédito con una Factura"""
    try:
        data = request.json
        nc_id = data.get('ncId')
        factura_id = data.get('facturaId')
        monto_compensar = float(data.get('montoCompensar', 0))
        
        sheet = get_sheet()
        ws = sheet.worksheet('Facturas')
        
        # Buscar NC
        cell_nc = ws.find(nc_id)
        if not cell_nc:
            return jsonify({'success': False, 'error': 'Nota de crédito no encontrada'}), 404
        
        # Buscar Factura
        cell_fac = ws.find(factura_id)
        if not cell_fac:
            return jsonify({'success': False, 'error': 'Factura no encontrada'}), 404
        
        row_nc = cell_nc.row
        row_fac = cell_fac.row
        
        # Obtener montos actuales
        monto_nc = float(ws.cell(row_nc, 10).value or 0)
        monto_fac = float(ws.cell(row_fac, 10).value or 0)
        
        # Si no se especifica monto, usar el menor de los dos
        if monto_compensar == 0:
            monto_compensar = min(abs(monto_nc), abs(monto_fac))
        
        fecha_hoy = datetime.now().strftime('%Y-%m-%d')
        
        # Actualizar NC
        ws.update_cell(row_nc, 12, 'TRUE')  # Pagado
        ws.update_cell(row_nc, 13, fecha_hoy)  # FechaPago
        ws.update_cell(row_nc, 18, factura_id)  # DocumentoRelacionado
        ws.update_cell(row_nc, 19, 'Compensado')  # Estado
        notas_nc = ws.cell(row_nc, 16).value or ''
        ws.update_cell(row_nc, 16, f"{notas_nc} | Compensado con FAC {factura_id}" if notas_nc else f"Compensado con FAC {factura_id}")
        
        # Actualizar Factura
        ws.update_cell(row_fac, 12, 'TRUE')  # Pagado
        ws.update_cell(row_fac, 13, fecha_hoy)  # FechaPago
        ws.update_cell(row_fac, 18, nc_id)  # DocumentoRelacionado
        ws.update_cell(row_fac, 19, 'Compensado')  # Estado
        notas_fac = ws.cell(row_fac, 16).value or ''
        ws.update_cell(row_fac, 16, f"{notas_fac} | Compensado con NC {nc_id}" if notas_fac else f"Compensado con NC {nc_id}")
        
        return jsonify({'success': True, 'montoCompensado': monto_compensar})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/facturas/pendientes/<cliente_id>', methods=['GET'])
@auth_required
def get_facturas_pendientes_cliente(cliente_id):
    """Obtiene facturas pendientes de un cliente para compensar con NC"""
    try:
        sheet = get_sheet()
        ws = sheet.worksheet('Facturas')
        records = ws.get_all_records()
        
        facturas = []
        for r in records:
            # Solo facturas (no NC), pendientes, del mismo cliente
            consecutivo = str(r.get('Consecutivo', ''))
            tipo_doc = r.get('TipoDocumento', '') or detectar_tipo_documento(consecutivo)
            estado = r.get('Estado', '')
            
            if not estado:
                estado = 'Pagado' if r.get('Pagado', 'FALSE') == 'TRUE' else 'Pendiente'
            
            if (tipo_doc == 'FAC' and 
                estado == 'Pendiente' and 
                str(r.get('ClienteID', '')) == cliente_id):
                facturas.append({
                    'id': str(r.get('ID', '')),
                    'consecutivo': consecutivo,
                    'fecha': r.get('Fecha', '')[:10] if r.get('Fecha') else '',
                    'montoCobrar': parse_number(r.get('MontoCobrar')),
                    'clienteNombre': r.get('ClienteNombre', '')
                })
        
        return jsonify({'success': True, 'data': facturas})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# =====================
# REPORTES
# =====================
def crear_estilo_excel():
    """Estilos para reportes Excel"""
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    return header_fill, header_font, border

def formato_moneda(valor):
    """Formatea número como moneda costarricense (sin símbolo especial para PDF)"""
    try:
        # Si ya es número, usarlo directamente
        if isinstance(valor, (int, float)):
            num = float(valor)
        else:
            # Convertir a string
            s = str(valor)
            # Eliminar ABSOLUTAMENTE TODO excepto dígitos, punto y guion
            limpio = ''
            tiene_punto = False
            for c in s:
                if c.isdigit():
                    limpio += c
                elif c == '.' and not tiene_punto:
                    limpio += c
                    tiene_punto = True
                elif c == '-' and len(limpio) == 0:
                    limpio += c
                # Ignorar TODO lo demás: comas, espacios, símbolos, unicode, etc.
            
            num = float(limpio) if limpio and limpio != '-' else 0.0
        
        # Formato con separadores de miles y 2 decimales
        return "{:,.2f}".format(num)
    except Exception as e:
        print(f"Error en formato_moneda: {valor} -> {e}")
        return "0.00"

def limpiar_texto(texto):
    """Limpia texto de caracteres especiales para PDF"""
    if not texto:
        return ""
    # Convertir a ASCII, reemplazando caracteres especiales
    resultado = ""
    for c in str(texto):
        if ord(c) < 128:  # Solo ASCII básico
            resultado += c
        elif c in 'áéíóúÁÉÍÓÚñÑüÜ':
            # Mantener acentos comunes del español
            resultado += c
        else:
            # Reemplazar otros caracteres por espacio
            resultado += ' '
    return resultado.strip()

def send_file_no_cache(buffer, mimetype, filename):
    """Envía archivo con headers anti-cache"""
    response = make_response(send_file(
        buffer,
        mimetype=mimetype,
        as_attachment=True,
        download_name=filename
    ))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

def crear_estado_cuenta_pdf(cliente_info, facturas_data, total_pendiente, empresa_config=None):
    """Genera un PDF de estado de cuenta presentable"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, 
                           leftMargin=0.75*inch, rightMargin=0.75*inch,
                           topMargin=0.5*inch, bottomMargin=0.75*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    hoy = datetime.now()
    
    # Configuración de empresa (usar valores por defecto si no hay config)
    if not empresa_config:
        empresa_config = {}
    
    emp_nombre = empresa_config.get('nombre', 'EMPRESA')
    emp_descripcion = empresa_config.get('descripcion', '')
    emp_telefono = empresa_config.get('telefono', '')
    emp_email = empresa_config.get('email', '')
    emp_mensaje = empresa_config.get('mensaje', 'Gracias por su preferencia')
    
    # === ENCABEZADO DE EMPRESA ===
    header_style = ParagraphStyle('Header', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#666666'))
    empresa_style = ParagraphStyle('Empresa', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1F4E79'), spaceAfter=2)
    
    elements.append(Paragraph(limpiar_texto(emp_nombre.upper()), empresa_style))
    if emp_descripcion:
        elements.append(Paragraph(limpiar_texto(emp_descripcion), header_style))
    
    contacto_parts = []
    if emp_telefono:
        contacto_parts.append(f"Tel: {emp_telefono}")
    if emp_email:
        contacto_parts.append(f"Email: {emp_email}")
    if contacto_parts:
        elements.append(Paragraph(" | ".join(contacto_parts), header_style))
    elements.append(Spacer(1, 20))
    
    # === TÍTULO ===
    titulo_style = ParagraphStyle('Titulo', parent=styles['Heading1'], 
                                   fontSize=16, alignment=TA_CENTER, 
                                   textColor=colors.HexColor('#1F4E79'),
                                   spaceAfter=20)
    elements.append(Paragraph("ESTADO DE CUENTA", titulo_style))
    
    # === INFO DEL CLIENTE ===
    cliente_box_data = [
        [Paragraph("<b>Cliente:</b>", styles['Normal']), Paragraph(limpiar_texto(cliente_info.get('nombre', '')), styles['Normal'])],
        [Paragraph("<b>Identificación:</b>", styles['Normal']), Paragraph(str(cliente_info.get('identificacion', '')), styles['Normal'])],
        [Paragraph("<b>Días de Crédito:</b>", styles['Normal']), Paragraph(str(cliente_info.get('diasCredito', 8)), styles['Normal'])],
        [Paragraph("<b>Fecha de Corte:</b>", styles['Normal']), Paragraph(hoy.strftime('%d/%m/%Y'), styles['Normal'])],
    ]
    cliente_table = Table(cliente_box_data, colWidths=[1.5*inch, 4*inch])
    cliente_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F5F5F5')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#DDDDDD')),
        ('PADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(cliente_table)
    elements.append(Spacer(1, 20))
    
    # === TABLA DE FACTURAS ===
    headers = ['Consecutivo', 'Fecha', 'Vencimiento', 'Total', 'CORFOGA', 'Saldo', 'Estado']
    table_data = [headers] + facturas_data
    
    col_widths = [1.4*inch, 0.8*inch, 0.8*inch, 0.9*inch, 0.7*inch, 0.9*inch, 0.9*inch]
    tabla = Table(table_data, colWidths=col_widths)
    
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (3, 1), (-2, -1), 'RIGHT'),  # Alinear montos a la derecha
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ])
    
    # Alternar colores
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F9F9F9'))
    
    tabla.setStyle(style)
    elements.append(tabla)
    elements.append(Spacer(1, 20))
    
    # === TOTAL ===
    total_style = ParagraphStyle('Total', parent=styles['Heading2'], 
                                  fontSize=14, alignment=TA_RIGHT,
                                  textColor=colors.HexColor('#1F4E79'))
    elements.append(Paragraph(f"<b>TOTAL PENDIENTE: CRC {formato_moneda(total_pendiente)}</b>", total_style))
    elements.append(Spacer(1, 30))
    
    # === PIE DE PÁGINA ===
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], 
                                   fontSize=9, textColor=colors.HexColor('#888888'),
                                   alignment=TA_CENTER)
    elements.append(Paragraph(limpiar_texto(emp_mensaje), footer_style))
    elements.append(Paragraph("Este documento es un estado de cuenta informativo", footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer

def crear_pdf_reporte(titulo, subtitulo, headers, datos, col_widths=None, resumen=None):
    """Genera un PDF con tabla de datos"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), 
                           leftMargin=0.5*inch, rightMargin=0.5*inch,
                           topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    titulo_style = ParagraphStyle('Titulo', parent=styles['Heading1'], 
                                   fontSize=16, alignment=TA_CENTER, spaceAfter=6)
    elements.append(Paragraph(titulo, titulo_style))
    
    # Subtítulo
    sub_style = ParagraphStyle('Subtitulo', parent=styles['Normal'], 
                                fontSize=10, alignment=TA_CENTER, textColor=colors.gray)
    elements.append(Paragraph(subtitulo, sub_style))
    elements.append(Spacer(1, 20))
    
    # Resumen si existe
    if resumen:
        for linea in resumen:
            res_style = ParagraphStyle('Resumen', parent=styles['Normal'], fontSize=11)
            elements.append(Paragraph(linea, res_style))
        elements.append(Spacer(1, 15))
    
    # Tabla
    table_data = [headers] + datos
    
    if col_widths:
        tabla = Table(table_data, colWidths=col_widths)
    else:
        tabla = Table(table_data)
    
    # Estilo de tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
    ])
    
    # Alternar colores de filas
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F5F5F5'))
    
    tabla.setStyle(style)
    elements.append(tabla)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer

# =====================
# REPORTES PDF
# =====================
@app.route('/api/reportes/semanal/pdf', methods=['GET'])
@auth_required
def reporte_semanal_pdf():
    """Genera reporte semanal en PDF"""
    try:
        sheet = get_sheet()
        ws = get_or_create_worksheet(sheet, 'Facturas', HEADERS_FACTURAS)
        
        try:
            facturas = ws.get_all_records()
        except:
            facturas = []
        
        # Obtener nombre de empresa de la configuración
        nombre_empresa = "EMPRESA"
        try:
            ws_config = sheet.worksheet('Configuracion')
            config_records = ws_config.get_all_records()
            for r in config_records:
                if r.get('Campo') == 'nombre':
                    nombre_empresa = r.get('Valor', 'EMPRESA')
                    break
        except:
            pass
        
        hoy = datetime.now()
        
        pendientes = [f for f in facturas if f.get('Pagado') != 'TRUE' and f.get('Estado') != 'Compensado']
        vencidas = []
        proximas = []
        
        for f in pendientes:
            try:
                fv_str = f.get('FechaVencimiento', '')
                if not fv_str:
                    continue
                fv = datetime.strptime(fv_str.split('T')[0], '%Y-%m-%d')
                dias = (fv - hoy).days
                if dias < 0:
                    vencidas.append({**f, 'diasAtraso': abs(dias)})
                elif dias <= 7:
                    proximas.append({**f, 'diasParaVencer': dias})
            except Exception as e:
                print(f"Error procesando fecha: {e}")
                pass
        
        total_pendiente = sum(parse_number(f.get('MontoCobrar')) for f in pendientes)
        total_vencido = sum(parse_number(f.get('MontoCobrar')) for f in vencidas)
        
        # Preparar datos para PDF
        headers = ['Consecutivo', 'Cliente', 'Fecha', 'Vencimiento', 'Dias', 'Monto (CRC)', 'Estado']
        datos = []
        
        # Vencidas primero
        for f in sorted(vencidas, key=lambda x: x.get('diasAtraso', 0), reverse=True):
            datos.append([
                str(f.get('Consecutivo', '')),
                limpiar_texto(str(f.get('ClienteNombre', ''))[:25]),
                f.get('Fecha', '')[:10],
                f.get('FechaVencimiento', '')[:10],
                str(f.get('diasAtraso', 0)),
                formato_moneda(parse_number(f.get('MontoCobrar'))),
                f"Vencida ({f.get('diasAtraso')}d)"
            ])
        
        # Próximas
        for f in sorted(proximas, key=lambda x: x.get('diasParaVencer', 0)):
            datos.append([
                str(f.get('Consecutivo', '')),
                limpiar_texto(str(f.get('ClienteNombre', ''))[:25]),
                f.get('Fecha', '')[:10],
                f.get('FechaVencimiento', '')[:10],
                str(f.get('diasParaVencer', 0)),
                formato_moneda(parse_number(f.get('MontoCobrar'))),
                f"Vence en {f.get('diasParaVencer')}d"
            ])
        
        resumen = [
            f"<b>Total por Cobrar:</b> CRC {formato_moneda(total_pendiente)}",
            f"<b>Facturas Pendientes:</b> {len(pendientes)}",
            f"<b>Facturas Vencidas:</b> {len(vencidas)} (CRC {formato_moneda(total_vencido)})",
        ]
        
        col_widths = [1.5*inch, 1.8*inch, 0.8*inch, 0.8*inch, 0.5*inch, 1.1*inch, 1*inch]
        
        buffer = crear_pdf_reporte(
            f"REPORTE SEMANAL CxC - {limpiar_texto(nombre_empresa)}",
            f"Generado: {hoy.strftime('%d/%m/%Y %H:%M')}",
            headers, datos, col_widths, resumen
        )
        
        return send_file_no_cache(buffer, 'application/pdf', f'Reporte_Semanal_{hoy.strftime("%Y%m%d")}.pdf')
    except Exception as e:
        print(f"Error en reporte_semanal_pdf: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reportes/cliente/<cliente_id>/pdf', methods=['GET'])
@auth_required
def reporte_cliente_pdf(cliente_id):
    """Estado de cuenta por cliente en PDF - Versión presentable"""
    try:
        sheet = get_sheet()
        ws_fac = sheet.worksheet('Facturas')
        ws_cli = sheet.worksheet('Clientes')
        
        # Obtener valores sin formato
        fac_values = ws_fac.get_all_values()
        cli_values = ws_cli.get_all_values()
        
        # Convertir a diccionarios manualmente
        fac_headers = fac_values[0] if fac_values else []
        facturas = []
        for row in fac_values[1:]:
            f = {}
            for i, h in enumerate(fac_headers):
                f[h] = row[i] if i < len(row) else ''
            facturas.append(f)
        
        cli_headers = cli_values[0] if cli_values else []
        clientes = []
        for row in cli_values[1:]:
            c = {}
            for i, h in enumerate(cli_headers):
                c[h] = row[i] if i < len(row) else ''
            clientes.append(c)
        
        cliente = None
        for c in clientes:
            if str(c.get('ID', '')) == cliente_id:
                cliente = c
                break
        
        if not cliente:
            return jsonify({'success': False, 'error': 'Cliente no encontrado'}), 404
        
        facturas_cliente = [f for f in facturas if str(f.get('ClienteID', '')) == cliente_id]
        hoy = datetime.now()
        
        pendientes = [f for f in facturas_cliente if f.get('Pagado') != 'TRUE']
        total_pendiente = sum(parse_number(f.get('MontoCobrar')) for f in pendientes)
        
        # Preparar datos de facturas
        facturas_data = []
        for f in sorted(pendientes, key=lambda x: x.get('FechaVencimiento', '')):
            try:
                fv = datetime.fromisoformat(f.get('FechaVencimiento', '').split('T')[0])
                dias = (fv - hoy).days
                estado = f"Vencida ({abs(dias)}d)" if dias < 0 else f"Vence en {dias}d"
            except:
                estado = "Pendiente"
            
            facturas_data.append([
                str(f.get('Consecutivo', '')),
                f.get('Fecha', '')[:10],
                f.get('FechaVencimiento', '')[:10],
                formato_moneda(parse_number(f.get('TotalFactura'))),
                formato_moneda(parse_number(f.get('CORFOGA'))),
                formato_moneda(parse_number(f.get('MontoCobrar'))),
                estado
            ])
        
        # Info del cliente
        cliente_info = {
            'nombre': cliente.get('Nombre', ''),
            'identificacion': cliente.get('Identificacion', ''),
            'diasCredito': cliente.get('DiasCredito', 8)
        }
        
        # Obtener configuración de empresa
        empresa_config = {}
        try:
            ws_config = sheet.worksheet('Configuracion')
            config_records = ws_config.get_all_records()
            for r in config_records:
                campo = r.get('Campo', '')
                valor = r.get('Valor', '')
                if campo:
                    empresa_config[campo] = valor
        except:
            pass  # Si no hay config, usar valores por defecto
        
        buffer = crear_estado_cuenta_pdf(cliente_info, facturas_data, total_pendiente, empresa_config)
        
        nombre = cliente.get('Nombre', 'Cliente')[:20].replace(' ', '_')
        return send_file_no_cache(buffer, 'application/pdf', f'Estado_Cuenta_{nombre}_{hoy.strftime("%Y%m%d")}.pdf')
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reportes/vencidas/pdf', methods=['GET'])
@auth_required
def reporte_vencidas_pdf():
    """Reporte de facturas vencidas en PDF"""
    try:
        sheet = get_sheet()
        ws = get_or_create_worksheet(sheet, 'Facturas', HEADERS_FACTURAS)
        
        try:
            facturas = ws.get_all_records()
        except:
            facturas = []
        
        # Obtener nombre de empresa
        nombre_empresa = "EMPRESA"
        try:
            ws_config = sheet.worksheet('Configuracion')
            config_records = ws_config.get_all_records()
            for r in config_records:
                if r.get('Campo') == 'nombre':
                    nombre_empresa = r.get('Valor', 'EMPRESA')
                    break
        except:
            pass
        
        hoy = datetime.now()
        vencidas = []
        
        for f in facturas:
            if f.get('Pagado') != 'TRUE' and f.get('Estado') != 'Compensado':
                try:
                    fv_str = f.get('FechaVencimiento', '')
                    if not fv_str:
                        continue
                    fv = datetime.strptime(fv_str.split('T')[0], '%Y-%m-%d')
                    dias = (hoy - fv).days
                    if dias > 0:
                        vencidas.append({**f, 'diasAtraso': dias})
                except:
                    pass
        
        vencidas.sort(key=lambda x: x.get('diasAtraso', 0), reverse=True)
        total_vencido = sum(parse_number(f.get('MontoCobrar')) for f in vencidas)
        
        headers = ['Consecutivo', 'Cliente', 'Cedula', 'Fecha', 'Vencimiento', 'Dias Atraso', 'Monto (CRC)']
        datos = []
        
        for f in vencidas:
            datos.append([
                str(f.get('Consecutivo', '')),
                limpiar_texto(str(f.get('ClienteNombre', ''))[:20]),
                str(f.get('CedulaCliente', '')),
                f.get('Fecha', '')[:10],
                f.get('FechaVencimiento', '')[:10],
                str(f.get('diasAtraso', 0)),
                formato_moneda(parse_number(f.get('MontoCobrar')))
            ])
        
        resumen = [
            f"<b>Total Facturas Vencidas:</b> {len(vencidas)}",
            f"<b>Monto Total Vencido:</b> CRC {formato_moneda(total_vencido)}",
        ]
        
        col_widths = [1.4*inch, 1.5*inch, 0.9*inch, 0.8*inch, 0.8*inch, 0.7*inch, 1*inch]
        
        buffer = crear_pdf_reporte(
            f"FACTURAS VENCIDAS - {limpiar_texto(nombre_empresa)}",
            f"Generado: {hoy.strftime('%d/%m/%Y %H:%M')}",
            headers, datos, col_widths, resumen
        )
        
        return send_file_no_cache(buffer, 'application/pdf', f'Facturas_Vencidas_{hoy.strftime("%Y%m%d")}.pdf')
    except Exception as e:
        print(f"Error en reporte_vencidas_pdf: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reportes/resumen-clientes/pdf', methods=['GET'])
@auth_required
def reporte_resumen_clientes_pdf():
    """Resumen por cliente en PDF"""
    try:
        sheet = get_sheet()
        ws_cli = get_or_create_worksheet(sheet, 'Clientes', HEADERS_CLIENTES)
        ws_fac = get_or_create_worksheet(sheet, 'Facturas', HEADERS_FACTURAS)
        
        try:
            clientes = ws_cli.get_all_records()
        except:
            clientes = []
        
        try:
            facturas = ws_fac.get_all_records()
        except:
            facturas = []
        
        hoy = datetime.now()
        resumen_data = []
        total_pendiente = 0
        total_vencido = 0
        
        for cliente in clientes:
            cliente_id = str(cliente.get('ID', ''))
            facs = [f for f in facturas if str(f.get('ClienteID', '')) == cliente_id]
            
            if not facs:
                continue
            
            pendientes = [f for f in facs if f.get('Pagado') != 'TRUE' and f.get('Estado') != 'Compensado']
            monto_pend = sum(parse_number(f.get('MontoCobrar')) for f in pendientes)
            
            monto_venc = 0
            for f in pendientes:
                try:
                    fv_str = f.get('FechaVencimiento', '')
                    if not fv_str:
                        continue
                    fv = datetime.strptime(fv_str.split('T')[0], '%Y-%m-%d')
                    if fv < hoy:
                        monto_venc += parse_number(f.get('MontoCobrar'))
                except:
                    pass
            
            if monto_pend > 0:
                resumen_data.append({
                    'nombre': cliente.get('Nombre', ''),
                    'identificacion': cliente.get('Identificacion', ''),
                    'pendientes': len(pendientes),
                    'monto': monto_pend,
                    'vencido': monto_venc
                })
                total_pendiente += monto_pend
                total_vencido += monto_venc
        
        resumen_data.sort(key=lambda x: x['monto'], reverse=True)
        
        headers = ['Cliente', 'Identificacion', 'Fact. Pend.', 'Pendiente (CRC)', 'Vencido']
        datos = []
        
        for r in resumen_data:
            datos.append([
                limpiar_texto(r['nombre'][:25]),
                str(r['identificacion']),
                str(r['pendientes']),
                formato_moneda(r['monto']),
                formato_moneda(r['vencido'])
            ])
        
        # Agregar fila de totales
        datos.append(['TOTAL', '', str(sum(r['pendientes'] for r in resumen_data)),
                     formato_moneda(total_pendiente), formato_moneda(total_vencido)])
        
        resumen = [
            f"<b>Total Clientes con Saldo:</b> {len(resumen_data)}",
            f"<b>Total por Cobrar:</b> CRC {formato_moneda(total_pendiente)}",
            f"<b>Total Vencido:</b> CRC {formato_moneda(total_vencido)}",
        ]
        
        col_widths = [2*inch, 1.2*inch, 0.8*inch, 1.3*inch, 1.3*inch]
        
        buffer = crear_pdf_reporte(
            "RESUMEN DE CUENTAS POR COBRAR",
            f"Fecha: {hoy.strftime('%d/%m/%Y')}",
            headers, datos, col_widths, resumen
        )
        
        return send_file_no_cache(buffer, 'application/pdf', f'Resumen_Clientes_{hoy.strftime("%Y%m%d")}.pdf')
    except Exception as e:
        print(f"Error en reporte_resumen_clientes_pdf: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# =====================
# REPORTES EXCEL
# =====================
@app.route('/api/reportes/semanal', methods=['GET'])
@auth_required
def reporte_semanal_excel():
    """Genera reporte semanal de CxC en Excel"""
    try:
        sheet = get_sheet()
        ws = get_or_create_worksheet(sheet, 'Facturas', HEADERS_FACTURAS)
        
        try:
            facturas = ws.get_all_records()
        except:
            facturas = []
        
        hoy = datetime.now()
        
        # Crear Excel
        wb = Workbook()
        ws_rep = wb.active
        ws_rep.title = "Reporte Semanal CxC"
        
        header_fill, header_font, border = crear_estilo_excel()
        
        # Título
        ws_rep.merge_cells('A1:H1')
        ws_rep['A1'] = f"REPORTE SEMANAL CxC - Gerald Ramírez"
        ws_rep['A1'].font = Font(bold=True, size=16)
        ws_rep['A2'] = f"Generado: {hoy.strftime('%d/%m/%Y %H:%M')}"
        ws_rep['A2'].font = Font(italic=True, size=10)
        
        # Resumen
        pendientes = [f for f in facturas if f.get('Pagado') != 'TRUE']
        vencidas = []
        proximas = []
        
        for f in pendientes:
            try:
                fv = datetime.fromisoformat(f.get('FechaVencimiento', '').split('T')[0])
                dias = (fv - hoy).days
                if dias < 0:
                    vencidas.append({**f, 'diasAtraso': abs(dias)})
                elif dias <= 7:
                    proximas.append({**f, 'diasParaVencer': dias})
            except:
                pass
        
        total_pendiente = sum(parse_number(f.get('MontoCobrar')) for f in pendientes)
        total_vencido = sum(parse_number(f.get('MontoCobrar')) for f in vencidas)
        
        ws_rep['A4'] = "RESUMEN"
        ws_rep['A4'].font = Font(bold=True, size=12)
        ws_rep['A5'] = f"Total por Cobrar: ₡{total_pendiente:,.2f}"
        ws_rep['A6'] = f"Facturas Pendientes: {len(pendientes)}"
        ws_rep['A7'] = f"Facturas Vencidas: {len(vencidas)}"
        ws_rep['A8'] = f"Monto Vencido: ₡{total_vencido:,.2f}"
        
        # Tabla de vencidas
        ws_rep['A10'] = "FACTURAS VENCIDAS"
        ws_rep['A10'].font = Font(bold=True, size=12, color="FF0000")
        
        headers = ['Consecutivo', 'Cliente', 'Fecha', 'Vencimiento', 'Días Atraso', 'Monto (CRC)', 'Tipo']
        for col, h in enumerate(headers, 1):
            cell = ws_rep.cell(row=11, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        for i, f in enumerate(sorted(vencidas, key=lambda x: x.get('diasAtraso', 0), reverse=True), 12):
            cell_cons = ws_rep.cell(row=i, column=1, value=str(f.get('Consecutivo', ''))); cell_cons.number_format = '@'; cell_cons.border = border
            ws_rep.cell(row=i, column=2, value=f.get('ClienteNombre', '')[:30]).border = border
            ws_rep.cell(row=i, column=3, value=f.get('Fecha', '')[:10]).border = border
            ws_rep.cell(row=i, column=4, value=f.get('FechaVencimiento', '')[:10]).border = border
            cell_dias = ws_rep.cell(row=i, column=5, value=f.get('diasAtraso', 0))
            cell_dias.border = border
            cell_dias.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            ws_rep.cell(row=i, column=6, value=parse_number(f.get('MontoCobrar'))).border = border
            ws_rep.cell(row=i, column=6).number_format = '₡#,##0.00'
            ws_rep.cell(row=i, column=7, value=f.get('TipoProducto', '')).border = border
        
        # Próximas a vencer
        row_start = 12 + len(vencidas) + 2
        ws_rep.cell(row=row_start, column=1, value="PRÓXIMAS A VENCER (7 días)")
        ws_rep.cell(row=row_start, column=1).font = Font(bold=True, size=12, color="FF8C00")
        
        for col, h in enumerate(headers, 1):
            cell = ws_rep.cell(row=row_start+1, column=col, value=h.replace('Días Atraso', 'Días para Vencer'))
            cell.fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
            cell.font = header_font
            cell.border = border
        
        for i, f in enumerate(sorted(proximas, key=lambda x: x.get('diasParaVencer', 0)), row_start+2):
            cell_cons = ws_rep.cell(row=i, column=1, value=str(f.get('Consecutivo', ''))); cell_cons.number_format = '@'; cell_cons.border = border
            ws_rep.cell(row=i, column=2, value=f.get('ClienteNombre', '')[:30]).border = border
            ws_rep.cell(row=i, column=3, value=f.get('Fecha', '')[:10]).border = border
            ws_rep.cell(row=i, column=4, value=f.get('FechaVencimiento', '')[:10]).border = border
            ws_rep.cell(row=i, column=5, value=f.get('diasParaVencer', 0)).border = border
            ws_rep.cell(row=i, column=6, value=parse_number(f.get('MontoCobrar'))).border = border
            ws_rep.cell(row=i, column=6).number_format = '₡#,##0.00'
            ws_rep.cell(row=i, column=7, value=f.get('TipoProducto', '')).border = border
        
        # Ajustar anchos
        for col in range(1, 8):
            ws_rep.column_dimensions[get_column_letter(col)].width = 18
        
        # Guardar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            attachment_filename=f'Reporte_Semanal_CxC_{hoy.strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reportes/cliente/<cliente_id>', methods=['GET'])
def reporte_cliente(cliente_id):
    """Estado de cuenta por cliente"""
    try:
        sheet = get_sheet()
        ws_fac = sheet.worksheet('Facturas')
        ws_cli = sheet.worksheet('Clientes')
        
        facturas = ws_fac.get_all_records()
        clientes = ws_cli.get_all_records()
        
        # Buscar cliente
        cliente = None
        for c in clientes:
            if str(c.get('ID', '')) == cliente_id:
                cliente = c
                break
        
        if not cliente:
            return jsonify({'success': False, 'error': 'Cliente no encontrado'}), 404
        
        # Filtrar facturas del cliente
        facturas_cliente = [f for f in facturas if str(f.get('ClienteID', '')) == cliente_id]
        
        hoy = datetime.now()
        
        # Crear Excel
        wb = Workbook()
        ws_rep = wb.active
        ws_rep.title = "Estado de Cuenta"
        
        header_fill, header_font, border = crear_estilo_excel()
        
        # Encabezado
        ws_rep.merge_cells('A1:G1')
        ws_rep['A1'] = "ESTADO DE CUENTA"
        ws_rep['A1'].font = Font(bold=True, size=16)
        
        ws_rep['A3'] = f"Cliente: {cliente.get('Nombre', '')}"
        ws_rep['A3'].font = Font(bold=True, size=12)
        ws_rep['A4'] = f"Cédula: {cliente.get('Identificacion', '')}"
        ws_rep['A5'] = f"Días de Crédito: {cliente.get('DiasCredito', 8)}"
        ws_rep['A6'] = f"Fecha: {hoy.strftime('%d/%m/%Y')}"
        
        # Resumen
        pendientes = [f for f in facturas_cliente if f.get('Pagado') != 'TRUE']
        pagadas = [f for f in facturas_cliente if f.get('Pagado') == 'TRUE']
        total_pendiente = sum(parse_number(f.get('MontoCobrar')) for f in pendientes)
        total_pagado = sum(parse_number(f.get('MontoCobrar')) for f in pagadas)
        
        ws_rep['A8'] = f"Total Pendiente: ₡{total_pendiente:,.2f}"
        ws_rep['A8'].font = Font(bold=True, color="FF0000")
        ws_rep['A9'] = f"Total Pagado: ₡{total_pagado:,.2f}"
        ws_rep['A9'].font = Font(bold=True, color="008000")
        
        # Tabla de facturas pendientes
        ws_rep['A11'] = "FACTURAS PENDIENTES"
        ws_rep['A11'].font = Font(bold=True, size=12)
        
        headers = ['Consecutivo', 'Fecha', 'Vencimiento', 'Total (CRC)', 'CORFOGA', 'Monto (CRC)', 'Estado']
        for col, h in enumerate(headers, 1):
            cell = ws_rep.cell(row=12, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        row = 13
        for f in sorted(pendientes, key=lambda x: x.get('FechaVencimiento', '')):
            try:
                fv = datetime.fromisoformat(f.get('FechaVencimiento', '').split('T')[0])
                dias = (fv - hoy).days
                estado = f"Vencida ({abs(dias)}d)" if dias < 0 else f"Vence en {dias}d"
            except:
                estado = "Pendiente"
            
            cell_cons = ws_rep.cell(row=row, column=1, value=str(f.get('Consecutivo', ''))); cell_cons.number_format = '@'; cell_cons.border = border
            ws_rep.cell(row=row, column=2, value=f.get('Fecha', '')[:10]).border = border
            ws_rep.cell(row=row, column=3, value=f.get('FechaVencimiento', '')[:10]).border = border
            ws_rep.cell(row=row, column=4, value=parse_number(f.get('TotalFactura'))).border = border
            ws_rep.cell(row=row, column=4).number_format = '₡#,##0.00'
            ws_rep.cell(row=row, column=5, value=parse_number(f.get('CORFOGA'))).border = border
            ws_rep.cell(row=row, column=5).number_format = '₡#,##0.00'
            ws_rep.cell(row=row, column=6, value=parse_number(f.get('MontoCobrar'))).border = border
            ws_rep.cell(row=row, column=6).number_format = '₡#,##0.00'
            ws_rep.cell(row=row, column=7, value=estado).border = border
            row += 1
        
        # Ajustar anchos
        ws_rep.column_dimensions['A'].width = 22  # Consecutivo más ancho
        for col in range(2, 8):
            ws_rep.column_dimensions[get_column_letter(col)].width = 16
        
        # Guardar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        nombre_cliente = cliente.get('Nombre', 'Cliente')[:20].replace(' ', '_')
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            attachment_filename=f'Estado_Cuenta_{nombre_cliente}_{hoy.strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reportes/vencidas', methods=['GET'])
def reporte_vencidas():
    """Reporte de facturas vencidas"""
    try:
        sheet = get_sheet()
        ws = sheet.worksheet('Facturas')
        facturas = ws.get_all_records()
        
        hoy = datetime.now()
        
        # Filtrar vencidas
        vencidas = []
        for f in facturas:
            if f.get('Pagado') == 'TRUE':
                continue
            try:
                fv = datetime.fromisoformat(f.get('FechaVencimiento', '').split('T')[0])
                dias = (fv - hoy).days
                if dias < 0:
                    vencidas.append({**f, 'diasAtraso': abs(dias)})
            except:
                pass
        
        # Crear Excel
        wb = Workbook()
        ws_rep = wb.active
        ws_rep.title = "Facturas Vencidas"
        
        header_fill, header_font, border = crear_estilo_excel()
        
        # Título
        ws_rep.merge_cells('A1:H1')
        ws_rep['A1'] = "REPORTE DE FACTURAS VENCIDAS"
        ws_rep['A1'].font = Font(bold=True, size=16, color="FF0000")
        ws_rep['A2'] = f"Generado: {hoy.strftime('%d/%m/%Y %H:%M')}"
        
        total_vencido = sum(parse_number(f.get('MontoCobrar')) for f in vencidas)
        ws_rep['A4'] = f"Total Facturas Vencidas: {len(vencidas)}"
        ws_rep['A5'] = f"Monto Total Vencido: ₡{total_vencido:,.2f}"
        ws_rep['A5'].font = Font(bold=True, color="FF0000")
        
        # Tabla
        headers = ['Consecutivo', 'Cliente', 'Cédula', 'Fecha', 'Vencimiento', 'Días Atraso', 'Monto (CRC)', 'Tipo']
        for col, h in enumerate(headers, 1):
            cell = ws_rep.cell(row=7, column=col, value=h)
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.font = header_font
            cell.border = border
        
        for i, f in enumerate(sorted(vencidas, key=lambda x: x.get('diasAtraso', 0), reverse=True), 8):
            cell_cons = ws_rep.cell(row=i, column=1, value=str(f.get('Consecutivo', ''))); cell_cons.number_format = '@'; cell_cons.border = border
            ws_rep.cell(row=i, column=2, value=f.get('ClienteNombre', '')[:25]).border = border
            ws_rep.cell(row=i, column=3, value=f.get('CedulaCliente', '')).border = border
            ws_rep.cell(row=i, column=4, value=f.get('Fecha', '')[:10]).border = border
            ws_rep.cell(row=i, column=5, value=f.get('FechaVencimiento', '')[:10]).border = border
            cell_dias = ws_rep.cell(row=i, column=6, value=f.get('diasAtraso', 0))
            cell_dias.border = border
            cell_dias.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            ws_rep.cell(row=i, column=7, value=parse_number(f.get('MontoCobrar'))).border = border
            ws_rep.cell(row=i, column=7).number_format = '₡#,##0.00'
            ws_rep.cell(row=i, column=8, value=f.get('TipoProducto', '')).border = border
        
        # Ajustar anchos
        for col in range(1, 9):
            ws_rep.column_dimensions[get_column_letter(col)].width = 16
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            attachment_filename=f'Facturas_Vencidas_{hoy.strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reportes/por-tipo', methods=['GET'])
@auth_required
def reporte_por_tipo():
    """Reporte agrupado por tipo de producto"""
    try:
        sheet = get_sheet()
        ws = sheet.worksheet('Facturas')
        facturas = ws.get_all_records()
        
        hoy = datetime.now()
        pendientes = [f for f in facturas if f.get('Pagado') != 'TRUE']
        
        # Agrupar por tipo
        por_tipo = {}
        for f in pendientes:
            tipo = f.get('TipoProducto', 'Sin especificar') or 'Sin especificar'
            if tipo not in por_tipo:
                por_tipo[tipo] = {'facturas': [], 'total': 0}
            por_tipo[tipo]['facturas'].append(f)
            por_tipo[tipo]['total'] += parse_number(f.get('MontoCobrar'))
        
        # Crear Excel
        wb = Workbook()
        ws_rep = wb.active
        ws_rep.title = "Por Tipo Producto"
        
        header_fill, header_font, border = crear_estilo_excel()
        
        ws_rep.merge_cells('A1:E1')
        ws_rep['A1'] = "REPORTE POR TIPO DE PRODUCTO"
        ws_rep['A1'].font = Font(bold=True, size=16)
        ws_rep['A2'] = f"Generado: {hoy.strftime('%d/%m/%Y %H:%M')}"
        
        # Resumen por tipo
        row = 4
        for tipo, data in sorted(por_tipo.items(), key=lambda x: x[1]['total'], reverse=True):
            ws_rep.cell(row=row, column=1, value=tipo).font = Font(bold=True)
            ws_rep.cell(row=row, column=2, value=f"{len(data['facturas'])} facturas")
            ws_rep.cell(row=row, column=3, value=data['total']).number_format = '₡#,##0.00'
            row += 1
        
        # Total general
        total_general = sum(d['total'] for d in por_tipo.values())
        row += 1
        ws_rep.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws_rep.cell(row=row, column=3, value=total_general).number_format = '₡#,##0.00'
        ws_rep.cell(row=row, column=3).font = Font(bold=True)
        
        for col in range(1, 4):
            ws_rep.column_dimensions[get_column_letter(col)].width = 25
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            attachment_filename=f'Reporte_Por_Tipo_{hoy.strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reportes/exportar-todo', methods=['GET'])
@auth_required
def exportar_todo():
    """Exporta todos los datos a Excel"""
    try:
        sheet = get_sheet()
        ws_fac = sheet.worksheet('Facturas')
        ws_cli = sheet.worksheet('Clientes')
        
        facturas = ws_fac.get_all_records()
        clientes = ws_cli.get_all_records()
        
        hoy = datetime.now()
        
        wb = Workbook()
        
        # Hoja de Facturas
        ws1 = wb.active
        ws1.title = "Facturas"
        
        header_fill, header_font, border = crear_estilo_excel()
        
        headers_fac = ['Consecutivo', 'Fecha', 'Cliente', 'Cédula', 'Tipo Producto', 'N° OC',
                       'Total (CRC)', 'CORFOGA', 'Otros Rebajos', 'Monto (CRC)',
                       'Vencimiento', 'Estado', 'Fecha Pago', 'Notas']
        
        for col, h in enumerate(headers_fac, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
        
        for i, f in enumerate(facturas, 2):
            pagado = f.get('Pagado') == 'TRUE'
            ws1.cell(row=i, column=1, value=str(f.get('Consecutivo', ''))).number_format = '@'
            ws1.cell(row=i, column=2, value=f.get('Fecha', '')[:10])
            ws1.cell(row=i, column=3, value=f.get('ClienteNombre', ''))
            ws1.cell(row=i, column=4, value=f.get('CedulaCliente', ''))
            ws1.cell(row=i, column=5, value=f.get('TipoProducto', ''))
            ws1.cell(row=i, column=6, value=f.get('OrdenCompra', ''))
            ws1.cell(row=i, column=7, value=parse_number(f.get('TotalFactura'))).number_format = '₡#,##0.00'
            ws1.cell(row=i, column=8, value=parse_number(f.get('CORFOGA'))).number_format = '₡#,##0.00'
            ws1.cell(row=i, column=9, value=parse_number(f.get('OtrosRebajos'))).number_format = '₡#,##0.00'
            ws1.cell(row=i, column=10, value=parse_number(f.get('MontoCobrar'))).number_format = '₡#,##0.00'
            ws1.cell(row=i, column=11, value=f.get('FechaVencimiento', '')[:10])
            ws1.cell(row=i, column=12, value='Pagada' if pagado else 'Pendiente')
            ws1.cell(row=i, column=13, value=f.get('FechaPago', '')[:10] if pagado else '')
            ws1.cell(row=i, column=14, value=f.get('Notas', ''))
        
        for col in range(1, 15):
            ws1.column_dimensions[get_column_letter(col)].width = 15
        
        # Hoja de Clientes
        ws2 = wb.create_sheet("Clientes")
        headers_cli = ['Nombre', 'Identificación', 'Días Crédito', 'Estado']
        
        for col, h in enumerate(headers_cli, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
        
        for i, c in enumerate(clientes, 2):
            ws2.cell(row=i, column=1, value=c.get('Nombre', ''))
            ws2.cell(row=i, column=2, value=c.get('Identificacion', ''))
            ws2.cell(row=i, column=3, value=c.get('DiasCredito', 8))
            ws2.cell(row=i, column=4, value='Activo' if c.get('Activo') == 'TRUE' else 'Inactivo')
        
        for col in range(1, 5):
            ws2.column_dimensions[get_column_letter(col)].width = 20
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            attachment_filename=f'Control_CxC_Gerald_{hoy.strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# =====================
# REPORTE: RESUMEN POR CLIENTE
# =====================
@app.route('/api/reportes/resumen-clientes', methods=['GET'])
def reporte_resumen_clientes():
    """Genera Excel con resumen de saldos por cliente"""
    try:
        sheet = get_sheet()
        ws_cli = get_or_create_worksheet(sheet, 'Clientes', HEADERS_CLIENTES)
        ws_fac = get_or_create_worksheet(sheet, 'Facturas', HEADERS_FACTURAS)
        
        clientes = ws_cli.get_all_records()
        facturas = ws_fac.get_all_records()
        
        hoy = datetime.now()
        
        # Calcular datos por cliente
        resumen = []
        total_facturas = 0
        total_pendiente = 0
        total_vencido = 0
        total_pagado_mes = 0
        
        inicio_mes = datetime(hoy.year, hoy.month, 1)
        
        for cliente in clientes:
            cliente_id = str(cliente.get('ID', ''))
            cliente_nombre = cliente.get('Nombre', '')
            
            # Filtrar facturas del cliente
            facs_cliente = [f for f in facturas if str(f.get('ClienteID', '')) == cliente_id]
            
            if not facs_cliente:
                continue
            
            # Calcular totales
            pendientes = [f for f in facs_cliente if f.get('Pagado') != 'TRUE']
            pagadas = [f for f in facs_cliente if f.get('Pagado') == 'TRUE']
            
            monto_pendiente = sum(parse_number(f.get('MontoCobrar')) for f in pendientes)
            
            # Calcular vencido
            monto_vencido = 0
            for f in pendientes:
                try:
                    fecha_venc = datetime.fromisoformat(f.get('FechaVencimiento', '')[:10])
                    if fecha_venc < hoy:
                        monto_vencido += parse_number(f.get('MontoCobrar'))
                except:
                    pass
            
            # Cobrado este mes
            cobrado_mes = 0
            for f in pagadas:
                try:
                    fecha_pago = f.get('FechaPago', '')
                    if fecha_pago:
                        fecha_pago_dt = datetime.fromisoformat(fecha_pago[:10])
                        if fecha_pago_dt >= inicio_mes:
                            cobrado_mes += parse_number(f.get('MontoCobrar'))
                except:
                    pass
            
            resumen.append({
                'nombre': cliente_nombre,
                'identificacion': cliente.get('Identificacion', ''),
                'total_facturas': len(facs_cliente),
                'facturas_pendientes': len(pendientes),
                'monto_pendiente': monto_pendiente,
                'monto_vencido': monto_vencido,
                'cobrado_mes': cobrado_mes
            })
            
            total_facturas += len(pendientes)
            total_pendiente += monto_pendiente
            total_vencido += monto_vencido
            total_pagado_mes += cobrado_mes
        
        # Ordenar por monto pendiente descendente
        resumen.sort(key=lambda x: x['monto_pendiente'], reverse=True)
        
        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen por Cliente"
        
        header_fill, header_font, border = crear_estilo_excel()
        danger_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        total_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        
        # Título
        ws.merge_cells('A1:G1')
        ws['A1'] = f"RESUMEN DE CUENTAS POR COBRAR - {hoy.strftime('%d/%m/%Y')}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Cliente', 'Identificación', 'Fact. Pend.', 'Pendiente (CRC)', 'Vencido', 'Cobrado (Mes)', '% Vencido']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        # Datos
        for i, r in enumerate(resumen, 4):
            ws.cell(row=i, column=1, value=r['nombre']).border = border
            ws.cell(row=i, column=2, value=r['identificacion']).border = border
            ws.cell(row=i, column=3, value=r['facturas_pendientes']).border = border
            
            cell_pend = ws.cell(row=i, column=4, value=r['monto_pendiente'])
            cell_pend.number_format = '₡#,##0.00'
            cell_pend.border = border
            
            cell_venc = ws.cell(row=i, column=5, value=r['monto_vencido'])
            cell_venc.number_format = '₡#,##0.00'
            cell_venc.border = border
            if r['monto_vencido'] > 0:
                cell_venc.fill = danger_fill
                cell_venc.font = Font(color="C62828", bold=True)
            
            cell_cob = ws.cell(row=i, column=6, value=r['cobrado_mes'])
            cell_cob.number_format = '₡#,##0.00'
            cell_cob.border = border
            
            # Porcentaje vencido
            pct = (r['monto_vencido'] / r['monto_pendiente'] * 100) if r['monto_pendiente'] > 0 else 0
            cell_pct = ws.cell(row=i, column=7, value=pct/100)
            cell_pct.number_format = '0%'
            cell_pct.border = border
            if pct > 50:
                cell_pct.fill = danger_fill
        
        # Fila de totales
        row_total = len(resumen) + 4
        ws.cell(row=row_total, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row_total, column=3, value=total_facturas).font = Font(bold=True)
        
        cell_tot_pend = ws.cell(row=row_total, column=4, value=total_pendiente)
        cell_tot_pend.number_format = '₡#,##0.00'
        cell_tot_pend.font = Font(bold=True)
        cell_tot_pend.fill = total_fill
        
        cell_tot_venc = ws.cell(row=row_total, column=5, value=total_vencido)
        cell_tot_venc.number_format = '₡#,##0.00'
        cell_tot_venc.font = Font(bold=True, color="C62828")
        
        cell_tot_cob = ws.cell(row=row_total, column=6, value=total_pagado_mes)
        cell_tot_cob.number_format = '₡#,##0.00'
        cell_tot_cob.font = Font(bold=True)
        
        for col in range(1, 8):
            ws.cell(row=row_total, column=col).border = border
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 12
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            attachment_filename=f'Resumen_Clientes_{hoy.strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# =====================
# HEALTH CHECK
# =====================
@app.route('/api/health', methods=['GET'])
def health_check():
    try:
        sheet = get_sheet()
        return jsonify({
            'success': True,
            'message': 'Conexión exitosa',
            'sheetName': sheet.title
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# =====================
# INICIAR
# =====================
# CONFIGURACIÓN
# =====================
HEADERS_CONFIG = ['Campo', 'Valor']

@app.route('/api/config', methods=['GET'])
@auth_required
def get_config():
    """Obtiene la configuración de la empresa"""
    try:
        sheet = get_sheet()
        try:
            ws = sheet.worksheet('Configuracion')
        except:
            # Crear hoja si no existe
            ws = sheet.add_worksheet(title='Configuracion', rows=20, cols=2)
            ws.append_row(HEADERS_CONFIG)
        
        records = ws.get_all_records()
        config = {}
        for r in records:
            campo = r.get('Campo', '')
            valor = r.get('Valor', '')
            if campo:
                config[campo] = valor
        
        return jsonify({'success': True, 'config': config})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/config', methods=['POST'])
@auth_required
def save_config():
    """Guarda la configuración de la empresa"""
    try:
        data = request.json
        sheet = get_sheet()
        
        try:
            ws = sheet.worksheet('Configuracion')
            # Obtener configuración existente para preservar campos extra
            existing_records = ws.get_all_records()
            existing_config = {}
            for r in existing_records:
                campo = r.get('Campo', '')
                valor = r.get('Valor', '')
                if campo:
                    existing_config[campo] = valor
        except:
            ws = sheet.add_worksheet(title='Configuracion', rows=20, cols=2)
            ws.append_row(HEADERS_CONFIG)
            existing_config = {}
        
        # Actualizar con los nuevos valores
        campos_guardar = ['nombre', 'cedula', 'descripcion', 'telefono', 'email', 'direccion', 'mensaje']
        for campo in campos_guardar:
            existing_config[campo] = data.get(campo, '')
        
        # Limpiar hoja y reescribir TODO (incluyendo campos extra como dashboardToken)
        ws.clear()
        ws.append_row(HEADERS_CONFIG)
        
        for campo, valor in existing_config.items():
            ws.append_row([campo, valor])
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# =====================
# PORTAL DE CLIENTES
# =====================
import hashlib
import secrets

def generar_token_cliente(cliente_id, negocio_id):
    """Genera un token único para el cliente"""
    data = f"{cliente_id}-{negocio_id}-portal"
    return hashlib.sha256(data.encode()).hexdigest()[:16]

def verificar_ultimos_digitos(identificacion, codigo):
    """Verifica los últimos 4 dígitos de la cédula"""
    # Limpiar identificación (quitar guiones, espacios)
    id_limpia = ''.join(filter(str.isdigit, str(identificacion)))
    return id_limpia[-4:] == codigo

@app.route('/api/portal/generar-link/<cliente_id>', methods=['POST'])
def generar_link_portal(cliente_id):
    """Genera un link de consulta para un cliente"""
    try:
        data = request.json or {}
        negocio_id = data.get('negocioId', 'default')
        
        token = generar_token_cliente(cliente_id, negocio_id)
        
        # Guardar el token en la hoja de clientes (opcional, para referencia)
        sheet = get_sheet()
        ws = sheet.worksheet('Clientes')
        records = ws.get_all_records()
        
        for i, r in enumerate(records):
            if str(r.get('ID', '')) == cliente_id:
                # Verificar si existe columna TokenPortal
                headers = ws.row_values(1)
                if 'TokenPortal' not in headers:
                    ws.update_cell(1, len(headers) + 1, 'TokenPortal')
                    headers.append('TokenPortal')
                
                col_token = headers.index('TokenPortal') + 1
                ws.update_cell(i + 2, col_token, token)
                break
        
        return jsonify({
            'success': True,
            'token': token,
            'link': f'/portal_clientes.html?t={token}'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/portal/info', methods=['GET'])
def portal_info():
    """Obtiene info básica del portal (nombre empresa)"""
    try:
        token = request.args.get('token')
        if not token:
            return jsonify({'success': False, 'error': 'Token requerido'}), 400
        
        sheet = get_sheet()
        
        # Buscar cliente con ese token
        ws_cli = sheet.worksheet('Clientes')
        clientes = ws_cli.get_all_records()
        
        cliente_encontrado = None
        for c in clientes:
            if c.get('TokenPortal') == token:
                cliente_encontrado = c
                break
        
        if not cliente_encontrado:
            return jsonify({'success': False, 'error': 'Link inválido o expirado'})
        
        # Obtener nombre de empresa
        empresa = ''
        try:
            ws_config = sheet.worksheet('Configuracion')
            config = ws_config.get_all_records()
            for r in config:
                if r.get('Campo') == 'nombre':
                    empresa = r.get('Valor', '')
                    break
        except:
            pass
        
        return jsonify({
            'success': True,
            'empresa': empresa,
            'clienteNombre': cliente_encontrado.get('Nombre', '')
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/portal/verificar', methods=['POST'])
def portal_verificar():
    """Verifica código y retorna datos del cliente"""
    try:
        data = request.json
        token = data.get('token')
        codigo = data.get('codigo')
        
        if not token or not codigo:
            return jsonify({'success': False, 'error': 'Datos incompletos'}), 400
        
        sheet = get_sheet()
        
        # Buscar cliente con ese token
        ws_cli = sheet.worksheet('Clientes')
        clientes = ws_cli.get_all_records()
        
        cliente = None
        for c in clientes:
            if c.get('TokenPortal') == token:
                cliente = c
                break
        
        if not cliente:
            return jsonify({'success': False, 'error': 'Link inválido'})
        
        # Verificar código (últimos 4 dígitos)
        if not verificar_ultimos_digitos(cliente.get('Identificacion', ''), codigo):
            return jsonify({'success': False, 'error': 'Código incorrecto'})
        
        # Obtener facturas del cliente
        ws_fac = get_or_create_worksheet(sheet, 'Facturas', HEADERS_FACTURAS)
        try:
            facturas = ws_fac.get_all_records()
        except:
            facturas = []
        
        cliente_id = str(cliente.get('ID', ''))
        
        # Separar pendientes y pagadas - usando set para evitar duplicados por ID
        hoy = datetime.now()
        pendientes = []
        pagos = []
        total_pendiente = 0
        total_vencido = 0
        ids_procesados = set()
        
        for f in facturas:
            # Evitar duplicados por ID
            factura_id = str(f.get('ID', ''))
            if factura_id in ids_procesados or not factura_id:
                continue
            ids_procesados.add(factura_id)
            
            # Solo facturas de este cliente
            if str(f.get('ClienteID', '')) != cliente_id:
                continue
            
            # Ignorar notas de crédito
            if f.get('TipoDocumento') == 'NC':
                continue
            
            pagado = str(f.get('Pagado', 'FALSE')).upper() == 'TRUE'
            estado = str(f.get('Estado', ''))
            
            if pagado or estado == 'Pagado' or estado == 'Compensado':
                # Factura pagada
                pagos.append({
                    'id': factura_id,
                    'consecutivo': str(f.get('Consecutivo', '')),
                    'fechaPago': f.get('FechaPago', ''),
                    'monto': parse_number(f.get('MontoCobrar', 0)),
                    'detalle': f.get('Notas', '') or f.get('DetallePago', '')
                })
            else:
                # Factura pendiente
                monto = parse_number(f.get('MontoCobrar', 0))
                total_abonado = parse_number(f.get('TotalAbonado', 0))
                saldo = monto - total_abonado
                
                if saldo <= 0:
                    continue
                
                total_pendiente += saldo
                
                try:
                    fv_str = str(f.get('FechaVencimiento', '')).split('T')[0]
                    if fv_str:
                        fv = datetime.strptime(fv_str, '%Y-%m-%d')
                        if fv < hoy:
                            total_vencido += saldo
                except:
                    pass
                
                pendientes.append({
                    'id': factura_id,
                    'consecutivo': str(f.get('Consecutivo', '')),
                    'fecha': str(f.get('Fecha', ''))[:10] if f.get('Fecha') else '',
                    'fechaVencimiento': str(f.get('FechaVencimiento', ''))[:10] if f.get('FechaVencimiento') else '',
                    'montoCobrar': saldo,
                    'totalAbonado': total_abonado
                })
        
        # Ordenar pendientes por fecha de vencimiento
        pendientes.sort(key=lambda x: x.get('fechaVencimiento', ''))
        
        # Ordenar pagos por fecha (más recientes primero)
        pagos.sort(key=lambda x: x.get('fechaPago', ''), reverse=True)
        pagos = pagos[:10]  # Solo últimos 10 pagos
        
        # Generar token de acceso temporal
        token_acceso = secrets.token_hex(8)
        
        return jsonify({
            'success': True,
            'tokenAcceso': token_acceso,
            'cliente': {
                'id': cliente_id,
                'nombre': cliente.get('Nombre', ''),
                'identificacion': cliente.get('Identificacion', ''),
                'diasCredito': cliente.get('DiasCredito', 8)
            },
            'facturas': pendientes,
            'pagos': pagos,
            'resumen': {
                'totalPendiente': total_pendiente,
                'totalVencido': total_vencido,
                'facturasPendientes': len(pendientes)
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/portal/estado-cuenta-pdf', methods=['GET'])
def portal_estado_cuenta_pdf():
    """Genera PDF de estado de cuenta para el portal"""
    try:
        token = request.args.get('token')
        
        if not token:
            return jsonify({'success': False, 'error': 'Token requerido'}), 400
        
        sheet = get_sheet()
        
        # Buscar cliente
        ws_cli = sheet.worksheet('Clientes')
        try:
            clientes = ws_cli.get_all_records()
        except Exception as e:
            print(f"Error leyendo clientes: {e}")
            return jsonify({'success': False, 'error': 'Error leyendo datos de clientes'}), 500
        
        cliente = None
        for c in clientes:
            # Buscar por TokenPortal o por ID si el token coincide con identificación
            if c.get('TokenPortal') == token or str(c.get('ID', '')) == token or str(c.get('Identificacion', '')) == token:
                cliente = c
                break
        
        if not cliente:
            return jsonify({'success': False, 'error': 'Token inválido o cliente no encontrado'}), 404
        
        cliente_id = str(cliente.get('ID', ''))
        
        # Obtener facturas
        ws_fac = sheet.worksheet('Facturas')
        facturas = ws_fac.get_all_records()
        
        hoy = datetime.now()
        facturas_cliente = [f for f in facturas if str(f.get('ClienteID', '')) == cliente_id and str(f.get('Pagado', '')).upper() != 'TRUE']
        
        total_pendiente = sum(parse_number(f.get('MontoCobrar', 0)) for f in facturas_cliente)
        
        # Preparar datos para PDF
        facturas_data = []
        for f in sorted(facturas_cliente, key=lambda x: x.get('FechaVencimiento', '')):
            try:
                fv = datetime.fromisoformat(f.get('FechaVencimiento', '').split('T')[0])
                dias = (fv - hoy).days
                estado = f"Vencida ({abs(dias)}d)" if dias < 0 else f"Vence en {dias}d"
            except:
                estado = "Pendiente"
            
            facturas_data.append([
                str(f.get('Consecutivo', '')),
                f.get('Fecha', '')[:10] if f.get('Fecha') else '',
                f.get('FechaVencimiento', '')[:10] if f.get('FechaVencimiento') else '',
                formato_moneda(parse_number(f.get('TotalFactura', 0))),
                formato_moneda(parse_number(f.get('CORFOGA', 0))),
                formato_moneda(parse_number(f.get('MontoCobrar', 0))),
                estado
            ])
        
        cliente_info = {
            'nombre': cliente.get('Nombre', ''),
            'identificacion': cliente.get('Identificacion', ''),
            'diasCredito': cliente.get('DiasCredito', 8)
        }
        
        # Obtener config empresa
        empresa_config = {}
        try:
            ws_config = sheet.worksheet('Configuracion')
            config = ws_config.get_all_records()
            for r in config:
                campo = r.get('Campo', '')
                if campo:
                    empresa_config[campo] = r.get('Valor', '')
        except:
            pass
        
        buffer = crear_estado_cuenta_pdf(cliente_info, facturas_data, total_pendiente, empresa_config)
        
        nombre = cliente.get('Nombre', 'Cliente')[:20].replace(' ', '_')
        return send_file_no_cache(buffer, 'application/pdf', f'Estado_Cuenta_{nombre}_{hoy.strftime("%Y%m%d")}.pdf')
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# =====================
# DASHBOARD SOLO LECTURA (para dueño del negocio)
# =====================

@app.route('/api/dashboard/generar-acceso', methods=['POST'])
@auth_required
def generar_acceso_dashboard():
    """Genera código de acceso para el dashboard de solo lectura"""
    try:
        data = request.json or {}
        codigo = data.get('codigo', '')
        
        if not codigo or len(codigo) < 4:
            return jsonify({'success': False, 'error': 'El código debe tener al menos 4 caracteres'}), 400
        
        # Generar token único
        token = secrets.token_hex(12)
        
        # Guardar en configuración
        sheet = get_sheet()
        try:
            ws = sheet.worksheet('Configuracion')
        except:
            ws = sheet.add_worksheet(title='Configuracion', rows=20, cols=2)
            ws.append_row(['Campo', 'Valor'])
        
        # Buscar si ya existe
        records = ws.get_all_records()
        found_token = False
        found_codigo = False
        
        for i, r in enumerate(records):
            if r.get('Campo') == 'dashboardToken':
                ws.update_cell(i + 2, 2, token)
                found_token = True
            if r.get('Campo') == 'dashboardCodigo':
                ws.update_cell(i + 2, 2, codigo)
                found_codigo = True
        
        if not found_token:
            ws.append_row(['dashboardToken', token])
        if not found_codigo:
            ws.append_row(['dashboardCodigo', codigo])
        
        return jsonify({
            'success': True,
            'token': token,
            'link': f'/dashboard_readonly.html?t={token}'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/dashboard/verificar', methods=['POST'])
def verificar_acceso_dashboard():
    """Verifica código de acceso al dashboard"""
    try:
        data = request.json
        token = data.get('token')
        codigo = data.get('codigo')
        
        if not token or not codigo:
            return jsonify({'success': False, 'error': 'Datos incompletos'}), 400
        
        sheet = get_sheet()
        
        # Verificar token y código
        try:
            ws_config = sheet.worksheet('Configuracion')
            config = ws_config.get_all_records()
        except:
            return jsonify({'success': False, 'error': 'No hay acceso configurado'})
        
        token_guardado = None
        codigo_guardado = None
        empresa = ''
        
        for r in config:
            campo = r.get('Campo', '')
            if campo == 'dashboardToken':
                token_guardado = r.get('Valor', '')
            elif campo == 'dashboardCodigo':
                codigo_guardado = r.get('Valor', '')
            elif campo == 'nombre':
                empresa = r.get('Valor', '')
        
        if token != token_guardado:
            return jsonify({'success': False, 'error': 'Link inválido'})
        
        if codigo != codigo_guardado:
            return jsonify({'success': False, 'error': 'Código incorrecto'})
        
        # Obtener todos los datos
        ws_cli = get_or_create_worksheet(sheet, 'Clientes', HEADERS_CLIENTES)
        ws_fac = get_or_create_worksheet(sheet, 'Facturas', HEADERS_FACTURAS)
        
        try:
            clientes_raw = ws_cli.get_all_records()
        except:
            clientes_raw = []
        
        try:
            facturas_raw = ws_fac.get_all_records()
        except:
            facturas_raw = []
        
        # Procesar clientes evitando duplicados
        clientes_ids = set()
        clientes = []
        for c in clientes_raw:
            cid = str(c.get('ID', ''))
            if cid and cid not in clientes_ids:
                clientes_ids.add(cid)
                clientes.append({
                    'id': cid,
                    'nombre': c.get('Nombre', ''),
                    'identificacion': str(c.get('Identificacion', ''))
                })
        
        # Procesar facturas evitando duplicados por ID
        facturas_ids = set()
        facturas = []
        for f in facturas_raw:
            fid = str(f.get('ID', ''))
            if fid and fid not in facturas_ids:
                facturas_ids.add(fid)
                facturas.append({
                    'id': fid,
                    'consecutivo': str(f.get('Consecutivo', '')),
                    'clienteId': str(f.get('ClienteID', '')),
                    'fecha': str(f.get('Fecha', ''))[:10] if f.get('Fecha') else '',
                    'fechaVencimiento': str(f.get('FechaVencimiento', ''))[:10] if f.get('FechaVencimiento') else '',
                    'totalFactura': parse_number(f.get('TotalFactura', 0)),
                    'montoCobrar': parse_number(f.get('MontoCobrar', 0)),
                    'pagado': str(f.get('Pagado', '')).upper() == 'TRUE' or f.get('Estado') == 'Pagado',
                    'fechaPago': f.get('FechaPago', ''),
                    'estado': f.get('Estado', ''),
                    'tipoDocumento': f.get('TipoDocumento', 'FAC')
                })
        
        return jsonify({
            'success': True,
            'tokenAcceso': secrets.token_hex(8),
            'empresa': empresa,
            'clientes': clientes,
            'facturas': facturas
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# =====================
# ABONOS PARCIALES
# =====================
@app.route('/api/abonos', methods=['GET'])
@auth_required
def get_abonos():
    """Obtiene todos los abonos, opcionalmente filtrados por factura"""
    try:
        factura_id = request.args.get('facturaId')
        sheet = get_sheet()
        ws = get_or_create_worksheet(sheet, 'Abonos', HEADERS_ABONOS)
        
        try:
            records = ws.get_all_records()
        except:
            records = []
        
        abonos = []
        for r in records:
            if factura_id and str(r.get('FacturaID', '')) != factura_id:
                continue
            abonos.append({
                'id': str(r.get('ID', '')),
                'facturaId': str(r.get('FacturaID', '')),
                'consecutivo': r.get('Consecutivo', ''),
                'fecha': r.get('Fecha', ''),
                'monto': parse_number(r.get('Monto', 0)),
                'metodoPago': r.get('MetodoPago', ''),
                'referencia': r.get('Referencia', ''),
                'notas': r.get('Notas', '')
            })
        
        return jsonify({'success': True, 'data': abonos})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/abonos', methods=['POST'])
@auth_required
def add_abono():
    """Registra un nuevo abono parcial"""
    try:
        data = request.json
        factura_id = data.get('facturaId')
        monto = float(data.get('monto', 0))
        
        if not factura_id or monto <= 0:
            return jsonify({'success': False, 'error': 'Factura y monto son requeridos'}), 400
        
        sheet = get_sheet()
        
        # Obtener la factura para validar
        ws_facturas = sheet.worksheet('Facturas')
        cell = ws_facturas.find(factura_id)
        if not cell:
            return jsonify({'success': False, 'error': 'Factura no encontrada'}), 404
        
        row = cell.row
        consecutivo_factura = ws_facturas.cell(row, 2).value  # Columna B = Consecutivo
        monto_cobrar = parse_number(ws_facturas.cell(row, 10).value)  # Columna J = MontoCobrar
        total_abonado_actual = parse_number(ws_facturas.cell(row, 20).value)  # Columna T = TotalAbonado
        
        # Calcular nuevo saldo
        nuevo_total_abonado = total_abonado_actual + monto
        nuevo_saldo = monto_cobrar - nuevo_total_abonado
        
        if nuevo_saldo < -0.01:  # Permitir pequeña tolerancia por redondeo
            return jsonify({
                'success': False, 
                'error': f'El abono excede el saldo pendiente. Saldo actual: {monto_cobrar - total_abonado_actual:.2f}'
            }), 400
        
        # Registrar el abono
        ws_abonos = get_or_create_worksheet(sheet, 'Abonos', HEADERS_ABONOS)
        abono_id = datetime.now().strftime('%Y%m%d%H%M%S%f')[:18]
        
        abono_row = [
            abono_id,
            factura_id,
            consecutivo_factura,
            data.get('fecha', datetime.now().strftime('%Y-%m-%d')),
            monto,
            data.get('metodoPago', 'Transferencia'),
            data.get('referencia', ''),
            data.get('notas', '')
        ]
        ws_abonos.append_row(abono_row)
        
        # Actualizar la factura con el nuevo total abonado y saldo
        ws_facturas.update_cell(row, 20, nuevo_total_abonado)  # TotalAbonado
        ws_facturas.update_cell(row, 21, max(0, nuevo_saldo))  # SaldoPendiente
        
        # Si el saldo es 0 o negativo, marcar como pagado
        if nuevo_saldo <= 0.01:
            ws_facturas.update_cell(row, 12, 'TRUE')  # Pagado
            ws_facturas.update_cell(row, 13, datetime.now().strftime('%Y-%m-%d'))  # FechaPago
            ws_facturas.update_cell(row, 19, 'Pagado')  # Estado
        else:
            ws_facturas.update_cell(row, 19, 'Abono Parcial')  # Estado
        
        return jsonify({
            'success': True,
            'abono': {
                'id': abono_id,
                'monto': monto,
                'totalAbonado': nuevo_total_abonado,
                'saldoPendiente': max(0, nuevo_saldo),
                'pagadoCompleto': nuevo_saldo <= 0.01
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/abonos/<abono_id>', methods=['DELETE'])
@auth_required
def delete_abono(abono_id):
    """Elimina un abono y recalcula el saldo de la factura"""
    try:
        sheet = get_sheet()
        ws_abonos = sheet.worksheet('Abonos')
        
        # Buscar el abono
        cell = ws_abonos.find(abono_id)
        if not cell:
            return jsonify({'success': False, 'error': 'Abono no encontrado'}), 404
        
        row = cell.row
        factura_id = ws_abonos.cell(row, 2).value  # FacturaID
        monto_abono = parse_number(ws_abonos.cell(row, 5).value)  # Monto
        
        # Eliminar el abono
        ws_abonos.delete_rows(row)
        
        # Recalcular totales de la factura
        ws_facturas = sheet.worksheet('Facturas')
        cell_factura = ws_facturas.find(factura_id)
        if cell_factura:
            row_factura = cell_factura.row
            monto_cobrar = parse_number(ws_facturas.cell(row_factura, 10).value)
            total_abonado_actual = parse_number(ws_facturas.cell(row_factura, 20).value)
            
            nuevo_total_abonado = max(0, total_abonado_actual - monto_abono)
            nuevo_saldo = monto_cobrar - nuevo_total_abonado
            
            ws_facturas.update_cell(row_factura, 20, nuevo_total_abonado)
            ws_facturas.update_cell(row_factura, 21, nuevo_saldo)
            
            # Actualizar estado
            if nuevo_total_abonado > 0:
                ws_facturas.update_cell(row_factura, 19, 'Abono Parcial')
                ws_facturas.update_cell(row_factura, 12, 'FALSE')
            else:
                ws_facturas.update_cell(row_factura, 19, 'Pendiente')
                ws_facturas.update_cell(row_factura, 12, 'FALSE')
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/facturas/<factura_id>/abonos', methods=['GET'])
@auth_required
def get_abonos_factura(factura_id):
    """Obtiene el historial de abonos de una factura específica"""
    try:
        sheet = get_sheet()
        ws_abonos = get_or_create_worksheet(sheet, 'Abonos', HEADERS_ABONOS)
        
        try:
            records = ws_abonos.get_all_records()
        except:
            records = []
        
        abonos = []
        total_abonado = 0
        for r in records:
            if str(r.get('FacturaID', '')) == factura_id:
                monto = parse_number(r.get('Monto', 0))
                total_abonado += monto
                abonos.append({
                    'id': str(r.get('ID', '')),
                    'fecha': r.get('Fecha', ''),
                    'monto': monto,
                    'metodoPago': r.get('MetodoPago', ''),
                    'referencia': r.get('Referencia', ''),
                    'notas': r.get('Notas', '')
                })
        
        # Ordenar por fecha descendente
        abonos.sort(key=lambda x: x['fecha'], reverse=True)
        
        return jsonify({
            'success': True,
            'data': abonos,
            'totalAbonado': total_abonado
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# =====================
# REPORTE ANTIGÜEDAD DE CARTERA
# =====================
@app.route('/api/reportes/antiguedad', methods=['GET'])
@auth_required
def get_antiguedad_cartera():
    """Genera reporte de antigüedad de cartera"""
    try:
        sheet = get_sheet()
        ws = get_or_create_worksheet(sheet, 'Facturas', HEADERS_FACTURAS)
        
        try:
            records = ws.get_all_records()
        except:
            records = []
        
        hoy = datetime.now().date()
        
        # Categorías de antigüedad
        cartera = {
            'corriente': {'facturas': [], 'total': 0, 'label': '0-30 días'},
            'dias_31_60': {'facturas': [], 'total': 0, 'label': '31-60 días'},
            'dias_61_90': {'facturas': [], 'total': 0, 'label': '61-90 días'},
            'dias_90_mas': {'facturas': [], 'total': 0, 'label': '+90 días'}
        }
        
        total_cartera = 0
        facturas_pendientes = 0
        
        # Set para evitar duplicados por ID
        ids_procesados = set()
        
        for r in records:
            # Obtener ID único de la factura
            factura_id = str(r.get('ID', ''))
            
            # Evitar duplicados - usar ID como identificador único
            if factura_id in ids_procesados or not factura_id:
                continue
            ids_procesados.add(factura_id)
            
            # Solo facturas pendientes (no pagadas, no NC)
            estado = str(r.get('Estado', ''))
            pagado = str(r.get('Pagado', 'FALSE')).upper()
            tipo_doc = str(r.get('TipoDocumento', 'FAC'))
            
            if pagado == 'TRUE' or estado == 'Pagado' or estado == 'Compensado' or tipo_doc == 'NC':
                continue
            
            # Calcular saldo pendiente
            monto_cobrar = parse_number(r.get('MontoCobrar', 0))
            total_abonado = parse_number(r.get('TotalAbonado', 0))
            saldo = monto_cobrar - total_abonado
            
            if saldo <= 0:
                continue
            
            # Calcular días de antigüedad desde fecha de vencimiento
            fecha_venc_str = str(r.get('FechaVencimiento', '')).split('T')[0] if r.get('FechaVencimiento') else ''
            dias_vencido = 0
            
            if fecha_venc_str:
                try:
                    fecha_venc = datetime.strptime(fecha_venc_str, '%Y-%m-%d').date()
                    dias_vencido = (hoy - fecha_venc).days
                except Exception as e:
                    print(f"Error parseando fecha {fecha_venc_str}: {e}")
                    dias_vencido = 0
            
            factura_info = {
                'id': factura_id,
                'consecutivo': str(r.get('Consecutivo', '')),
                'clienteNombre': r.get('ClienteNombre', ''),
                'clienteId': str(r.get('ClienteID', '')),
                'fecha': str(r.get('Fecha', ''))[:10] if r.get('Fecha') else '',
                'fechaVencimiento': fecha_venc_str,
                'montoCobrar': monto_cobrar,
                'totalAbonado': total_abonado,
                'saldoPendiente': saldo,
                'diasVencido': max(0, dias_vencido)
            }
            
            total_cartera += saldo
            facturas_pendientes += 1
            
            # Clasificar por antigüedad (días desde vencimiento)
            if dias_vencido <= 30:
                cartera['corriente']['facturas'].append(factura_info)
                cartera['corriente']['total'] += saldo
            elif dias_vencido <= 60:
                cartera['dias_31_60']['facturas'].append(factura_info)
                cartera['dias_31_60']['total'] += saldo
            elif dias_vencido <= 90:
                cartera['dias_61_90']['facturas'].append(factura_info)
                cartera['dias_61_90']['total'] += saldo
            else:
                cartera['dias_90_mas']['facturas'].append(factura_info)
                cartera['dias_90_mas']['total'] += saldo
        
        # Calcular porcentajes
        for key in cartera:
            if total_cartera > 0:
                cartera[key]['porcentaje'] = round((cartera[key]['total'] / total_cartera) * 100, 2)
            else:
                cartera[key]['porcentaje'] = 0
            cartera[key]['cantidad'] = len(cartera[key]['facturas'])
            # Ordenar facturas por días vencido (más vencidas primero)
            cartera[key]['facturas'].sort(key=lambda x: x['diasVencido'], reverse=True)
        
        return jsonify({
            'success': True,
            'data': {
                'cartera': cartera,
                'resumen': {
                    'totalCartera': total_cartera,
                    'facturasPendientes': facturas_pendientes,
                    'fechaReporte': hoy.strftime('%Y-%m-%d')
                }
            }
        })
    except Exception as e:
        print(f"Error en get_antiguedad_cartera: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reportes/antiguedad/excel', methods=['GET'])
@auth_required
def export_antiguedad_excel():
    """Exporta reporte de antigüedad a Excel"""
    try:
        # Obtener datos
        sheet = get_sheet()
        ws = get_or_create_worksheet(sheet, 'Facturas', HEADERS_FACTURAS)
        
        try:
            records = ws.get_all_records()
        except:
            records = []
        
        hoy = datetime.now().date()
        
        # Crear workbook
        wb = Workbook()
        ws_excel = wb.active
        ws_excel.title = "Antigüedad de Cartera"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        corriente_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        dias_31_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        dias_61_fill = PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid")
        dias_90_fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws_excel['A1'] = "REPORTE DE ANTIGÜEDAD DE CARTERA"
        ws_excel['A1'].font = Font(bold=True, size=14)
        ws_excel['A2'] = f"Fecha: {hoy.strftime('%d/%m/%Y')}"
        
        # Headers
        headers = ['Consecutivo', 'Cliente', 'Fecha', 'Vencimiento', 'Monto', 'Abonado', 'Saldo', 'Días Vencido', 'Categoría']
        for col, header in enumerate(headers, 1):
            cell = ws_excel.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        row_num = 5
        for r in records:
            estado = r.get('Estado', '')
            pagado = r.get('Pagado', 'FALSE')
            tipo_doc = r.get('TipoDocumento', 'FAC')
            
            if pagado == 'TRUE' or estado == 'Pagado' or estado == 'Compensado' or tipo_doc == 'NC':
                continue
            
            monto_cobrar = parse_number(r.get('MontoCobrar', 0))
            total_abonado = parse_number(r.get('TotalAbonado', 0))
            saldo = monto_cobrar - total_abonado
            
            if saldo <= 0:
                continue
            
            fecha_venc_str = r.get('FechaVencimiento', '')
            dias_vencido = 0
            if fecha_venc_str:
                try:
                    fecha_venc = datetime.strptime(fecha_venc_str, '%Y-%m-%d').date()
                    dias_vencido = max(0, (hoy - fecha_venc).days)
                except:
                    pass
            
            # Determinar categoría
            if dias_vencido <= 30:
                categoria = '0-30 días'
                fill = corriente_fill
            elif dias_vencido <= 60:
                categoria = '31-60 días'
                fill = dias_31_fill
            elif dias_vencido <= 90:
                categoria = '61-90 días'
                fill = dias_61_fill
            else:
                categoria = '+90 días'
                fill = dias_90_fill
            
            data_row = [
                r.get('Consecutivo', ''),
                r.get('ClienteNombre', ''),
                r.get('Fecha', ''),
                fecha_venc_str,
                monto_cobrar,
                total_abonado,
                saldo,
                dias_vencido,
                categoria
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = ws_excel.cell(row=row_num, column=col, value=value)
                cell.border = border
                if col == 9:  # Columna categoría
                    cell.fill = fill
                if col in [5, 6, 7]:  # Columnas monetarias
                    cell.number_format = '#,##0.00'
            
            row_num += 1
        
        # Ajustar anchos
        column_widths = [18, 30, 12, 12, 15, 15, 15, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws_excel.column_dimensions[get_column_letter(i)].width = width
        
        # Guardar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=antiguedad_cartera_{hoy.strftime("%Y%m%d")}.xlsx'
        
        return response
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reportes/antiguedad/pdf', methods=['GET'])
@auth_required
def export_antiguedad_pdf():
    """Exporta reporte de antigüedad a PDF"""
    try:
        sheet = get_sheet()
        ws = get_or_create_worksheet(sheet, 'Facturas', HEADERS_FACTURAS)
        
        try:
            records = ws.get_all_records()
        except:
            records = []
        
        hoy = datetime.now().date()
        
        # Crear PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), 
                               leftMargin=0.5*inch, rightMargin=0.5*inch,
                               topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            alignment=TA_CENTER,
            spaceAfter=20
        )
        elements.append(Paragraph("REPORTE DE ANTIGÜEDAD DE CARTERA", title_style))
        elements.append(Paragraph(f"Fecha: {hoy.strftime('%d/%m/%Y')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Preparar datos para tabla
        table_data = [['Consecutivo', 'Cliente', 'Vencimiento', 'Monto', 'Abonado', 'Saldo', 'Días', 'Categoría']]
        
        totales = {'corriente': 0, 'dias_31': 0, 'dias_61': 0, 'dias_90': 0}
        
        for r in records:
            estado = r.get('Estado', '')
            pagado = r.get('Pagado', 'FALSE')
            tipo_doc = r.get('TipoDocumento', 'FAC')
            
            if pagado == 'TRUE' or estado == 'Pagado' or estado == 'Compensado' or tipo_doc == 'NC':
                continue
            
            monto_cobrar = parse_number(r.get('MontoCobrar', 0))
            total_abonado = parse_number(r.get('TotalAbonado', 0))
            saldo = monto_cobrar - total_abonado
            
            if saldo <= 0:
                continue
            
            fecha_venc_str = r.get('FechaVencimiento', '')
            dias_vencido = 0
            if fecha_venc_str:
                try:
                    fecha_venc = datetime.strptime(fecha_venc_str, '%Y-%m-%d').date()
                    dias_vencido = max(0, (hoy - fecha_venc).days)
                except:
                    pass
            
            if dias_vencido <= 30:
                categoria = '0-30'
                totales['corriente'] += saldo
            elif dias_vencido <= 60:
                categoria = '31-60'
                totales['dias_31'] += saldo
            elif dias_vencido <= 90:
                categoria = '61-90'
                totales['dias_61'] += saldo
            else:
                categoria = '+90'
                totales['dias_90'] += saldo
            
            table_data.append([
                r.get('Consecutivo', '')[:20],
                limpiar_texto(r.get('ClienteNombre', '')[:25]),
                fecha_venc_str,
                formato_moneda(monto_cobrar),
                formato_moneda(total_abonado),
                formato_moneda(saldo),
                str(dias_vencido),
                categoria
            ])
        
        # Crear tabla
        if len(table_data) > 1:
            table = Table(table_data, colWidths=[1.3*inch, 1.8*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.5*inch, 0.6*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')])
            ]))
            elements.append(table)
        
        # Resumen
        elements.append(Spacer(1, 20))
        total_cartera = sum(totales.values())
        resumen_data = [
            ['RESUMEN', 'Monto (CRC)', '%'],
            ['0-30 dias', formato_moneda(totales['corriente']), f"{(totales['corriente']/total_cartera*100) if total_cartera > 0 else 0:.1f}%"],
            ['31-60 dias', formato_moneda(totales['dias_31']), f"{(totales['dias_31']/total_cartera*100) if total_cartera > 0 else 0:.1f}%"],
            ['61-90 dias', formato_moneda(totales['dias_61']), f"{(totales['dias_61']/total_cartera*100) if total_cartera > 0 else 0:.1f}%"],
            ['+90 dias', formato_moneda(totales['dias_90']), f"{(totales['dias_90']/total_cartera*100) if total_cartera > 0 else 0:.1f}%"],
            ['TOTAL', formato_moneda(total_cartera), '100%']
        ]
        
        resumen_table = Table(resumen_data, colWidths=[1.5*inch, 1.5*inch, 1*inch])
        resumen_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E5E7EB')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        elements.append(resumen_table)
        
        doc.build(elements)
        buffer.seek(0)
        
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=antiguedad_cartera_{hoy.strftime("%Y%m%d")}.pdf'
        
        return response
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# =====================
# DASHBOARD STATS
# =====================
@app.route('/api/dashboard/stats', methods=['GET'])
@auth_required
def get_dashboard_stats():
    """Obtiene estadísticas para el dashboard"""
    try:
        sheet = get_sheet()
        ws = get_or_create_worksheet(sheet, 'Facturas', HEADERS_FACTURAS)
        
        try:
            records = ws.get_all_records()
        except:
            records = []
        
        hoy = datetime.now().date()
        
        stats = {
            'totalPorCobrar': 0,
            'totalCobrado': 0,
            'facturasPendientes': 0,
            'facturasPagadas': 0,
            'facturasVencidas': 0,
            'proximasVencer': 0,  # Próximos 7 días
            'clientesConDeuda': set(),
            'porCategoria': {
                'corriente': 0,
                'dias_31_60': 0,
                'dias_61_90': 0,
                'dias_90_mas': 0
            }
        }
        
        for r in records:
            tipo_doc = r.get('TipoDocumento', 'FAC')
            if tipo_doc == 'NC':
                continue
                
            monto_cobrar = parse_number(r.get('MontoCobrar', 0))
            total_abonado = parse_number(r.get('TotalAbonado', 0))
            saldo = monto_cobrar - total_abonado
            
            estado = r.get('Estado', '')
            pagado = r.get('Pagado', 'FALSE')
            
            if pagado == 'TRUE' or estado == 'Pagado' or estado == 'Compensado':
                stats['totalCobrado'] += monto_cobrar
                stats['facturasPagadas'] += 1
            else:
                if saldo > 0:
                    stats['totalPorCobrar'] += saldo
                    stats['facturasPendientes'] += 1
                    stats['clientesConDeuda'].add(r.get('ClienteNombre', ''))
                    
                    # Verificar vencimiento
                    fecha_venc_str = r.get('FechaVencimiento', '')
                    if fecha_venc_str:
                        try:
                            fecha_venc = datetime.strptime(fecha_venc_str, '%Y-%m-%d').date()
                            dias_para_vencer = (fecha_venc - hoy).days
                            dias_vencido = (hoy - fecha_venc).days
                            
                            if dias_para_vencer < 0:
                                stats['facturasVencidas'] += 1
                            elif dias_para_vencer <= 7:
                                stats['proximasVencer'] += 1
                            
                            # Clasificar
                            if dias_vencido <= 30:
                                stats['porCategoria']['corriente'] += saldo
                            elif dias_vencido <= 60:
                                stats['porCategoria']['dias_31_60'] += saldo
                            elif dias_vencido <= 90:
                                stats['porCategoria']['dias_61_90'] += saldo
                            else:
                                stats['porCategoria']['dias_90_mas'] += saldo
                        except:
                            stats['porCategoria']['corriente'] += saldo
        
        stats['clientesConDeuda'] = len(stats['clientesConDeuda'])
        
        return jsonify({'success': True, 'data': stats})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# =====================
# INICIAR SERVIDOR
# =====================
if __name__ == '__main__':
    print("="*50)
    print("Control CxC v4 - Multi-cliente")
    print("="*50)
    print(f"Credenciales: {CREDENTIALS_FILE}")
    print(f"Negocios: {NEGOCIOS_FILE}")
    negocios = load_negocios()
    print(f"Negocios configurados: {len(negocios)}")
    print("-"*50)
    print("Servidor: http://localhost:5000")
    print("="*50)
    app.run(debug=True, port=5000)
